"""Step E — extend Investments, extend Checks, extend Dashboard, set sheet
order and tab colours."""
import sys, datetime as dt
sys.path.insert(0, "build")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import style as S

F = "work.xlsx"
wb = openpyxl.load_workbook(F)
FX_INR = "'FX & Assumptions'!$B$7"
FX_USD = "'FX & Assumptions'!$B$9"

# ========================================================== INVESTMENTS =====
ws = wb["Investments"]
S.widths(ws, {"K": 18, "L": 13})
for r in (1, 2):
    for col in ("K", "L"):
        ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=S.TITLE_BG)
for col in ("K", "L"):
    ws[f"{col}3"].fill = PatternFill("solid", fgColor=S.HEAD_BG)
    for r in (14, 20):
        ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=S.BAND_BG)
S.header(ws, 5, ["Value (AED)", "Weight"], col0=11)
for r in range(6, 12):
    S.put(ws, f"K{r}", f"=E{r}*{FX_INR}", fmt=S.AED)
    S.put(ws, f"L{r}", f"=IF($E$12=0,0,E{r}/$E$12)", fmt=S.PCT)
S.put(ws, "K12", f"=E12*{FX_INR}", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "L12", "=IF($E$12=0,0,E12/$E$12)", fmt=S.PCT, bold=True, fill=S.HEAD_BG)

W = 12
S.band(ws, 31, "CONSOLIDATED PORTFOLIO IN DIRHAMS", W)
S.header(ws, 32, ["Holding group", "Native value", "Currency", "Rate to AED", "Value (AED)",
                  "Share", "", "", "", "Note", "", ""])
cons = [
    (33, "Indian mutual funds — 6 schemes", "=E12", "INR", f"={FX_INR}",
     "Five Nippon schemes plus one Motilal Oswal midcap fund, valued at the 10 Aug MF Central "
     "snapshot."),
    (34, "ICICI settlement cash", "='Inputs'!B40", "INR", f"={FX_INR}",
     "Rupee cash left idle after the August SIPs."),
    (35, "Amana trading account", "=B22", "USD", f"={FX_USD}",
     "Ending account value across 41 open positions."),
]
for r, label, native, ccy, rate, note in cons:
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", native, fmt={"INR": S.INR, "USD": S.USD}[ccy], color=S.LINK_GREEN)
    S.put(ws, f"C{r}", ccy)
    S.put(ws, f"D{r}", rate, fmt=S.NUM4, color=S.LINK_GREEN)
    S.put(ws, f"E{r}", f"=B{r}*D{r}", fmt=S.AED)
    S.put(ws, f"F{r}", f"=IF($E$36=0,0,E{r}/$E$36)", fmt=S.PCT)
    S.put(ws, f"J{r}", note, wrap=True)
tot = [
    (36, "TOTAL INVESTED (AED)", "=SUM(E33:E35)", S.AED,
     "This is the figure the Net Worth and Wealth Plan sheets grow."),
    (37, "Total cost basis (AED)", f"=D12*{FX_INR}+B23*{FX_USD}+'Inputs'!B40*{FX_INR}", S.AED,
     "Mutual-fund cost plus net money paid into Amana plus the idle broker cash. The cash is "
     "carried at cost so that holding it never shows up as a gain."),
    (38, "Total gain / (loss) (AED)", "=E36-E37", S.AED,
     "Snapshot gain, not a forecast. The Wealth Plan deliberately assumes less."),
    (39, "Return on cost", "=IF(E37=0,0,E38/E37)", S.PCT,
     "Simple return on money in. Not annualised — most of this capital has been invested for less "
     "than a full year."),
]
for r, label, formula, fmt, note in tot:
    S.put(ws, f"A{r}", label, bold=True, fill=S.HEAD_BG)
    for col in "BCD":
        S.put(ws, f"{col}{r}", None, fill=S.HEAD_BG)
    S.put(ws, f"E{r}", formula, fmt=fmt, bold=True, fill=S.HEAD_BG)
    S.put(ws, f"F{r}", None, fill=S.HEAD_BG)
    S.put(ws, f"J{r}", note, wrap=True, bold=True, fill=S.HEAD_BG)

S.band(ws, 41, "ASSET ALLOCATION — ACTUAL AGAINST TARGET", W)
S.header(ws, 42, ["Asset class", "Value (AED)", "Actual weight", "Target weight", "Drift",
                  "Action", "", "", "", "Note", "", ""])
alloc = [
    (43, "Indian large cap", f"=E6*{FX_INR}", 0.30,
     "Nippon Large Cap. The stabiliser of the portfolio."),
    (44, "Indian multi cap", f"=E7*{FX_INR}", 0.20,
     "Nippon Multi Cap. Already close to target."),
    (45, "Indian mid cap", f"=(E8+E11)*{FX_INR}", 0.15,
     "Nippon Growth Mid Cap plus Motilal Oswal Midcap. Two funds doing the same job — a candidate "
     "for consolidation."),
    (46, "Indian small cap", f"=E9*{FX_INR}", 0.10,
     "Nippon Small Cap. The highest-volatility sleeve; keep it small and never sell it in a panic."),
    (47, "Commodity — silver", f"=E10*{FX_INR}", 0.05,
     "Nippon Silver ETF FoF. The SIP is cancelled but the holding remains; it is the best-performing "
     "line in the portfolio and worth keeping as a diversifier."),
    (48, "Global equity", f"=B22*{FX_USD}", 0.15,
     "Amana account. The only non-rupee investment exposure in the file."),
    (49, "Broker cash", f"='Inputs'!B40*{FX_INR}", 0.05,
     "Idle settlement cash. Target is really zero; 5% is the practical tolerance."),
]
for r, label, val, target, note in alloc:
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", val, fmt=S.AED)
    S.put(ws, f"C{r}", "=IF($B$50=0,0,B%d/$B$50)" % r, fmt=S.PCT)
    S.put(ws, f"D{r}", target, fmt=S.PCT, color=S.INPUT_BLUE, fill=S.LEVER_FILL)
    S.put(ws, f"E{r}", f"=C{r}-D{r}", fmt=S.PCT)
    S.put(ws, f"F{r}", f'=IF(ABS(E{r})<=0.05,"On target",IF(E{r}>0,"Overweight","Underweight"))')
    S.put(ws, f"J{r}", note, wrap=True)
S.put(ws, "A50", "TOTAL", bold=True, fill=S.HEAD_BG)
S.put(ws, "B50", "=SUM(B43:B49)", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "C50", "=IF(B50=0,0,SUM(B43:B49)/B50)", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, "D50", "=SUM(D43:D49)", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, "E50", "=C50-D50", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, "F50", '=IF(ABS(D50-1)<0.0001,"Targets valid","TARGETS DO NOT SUM TO 100%")',
      bold=True, fill=S.HEAD_BG)
S.put(ws, "J50", "Rebalance by directing new SIP instalments, never by selling. Selling triggers "
                 "exit loads and capital gains; redirecting costs nothing.",
      wrap=True, bold=True, fill=S.HEAD_BG)

S.band(ws, 52, "RISK CONTROLS", W)
S.header(ws, 53, ["Control", "Value", "Limit", "Status", "", "", "", "", "", "Read", "", ""])
risk = [
    (54, "Single fund house — Nippon", "=B16", 0.80, S.PCT,
     "Five of six mutual-fund holdings sit with one asset management company. Market risk is "
     "diversified; operational and manager risk is not."),
    (55, "Largest single fund", "=MAX(L6:L11)", 0.35, S.PCT,
     "No single fund should dominate. Concentration here is what turns one bad manager into a bad "
     "portfolio."),
    (56, "Small and mid cap combined", "=IF(B50=0,0,(B45+B46)/B50)", 0.30, S.PCT,
     "The volatile end. Fine at this size for a long horizon; painful if it has to be sold early."),
    (57, "Commodity sleeve", "=IF(B50=0,0,B47/B50)", 0.15, S.PCT,
     "Silver diversifies equity but produces no income. A modest allocation is the right size."),
    (58, "Rupee currency exposure", "=IF(B50=0,0,(B43+B44+B45+B46+B47+B49)/B50)", 0.85, S.PCT,
     "Almost the entire portfolio is denominated in rupees while every liability is in dirhams. "
     "The dirham is pegged to the dollar, so this is a genuine, unhedged currency mismatch."),
    (59, "Dollar currency exposure", "=IF(B50=0,0,B48/B50)", 0.50, S.PCT,
     "Effectively dirham-linked because of the peg, which makes it the natural home for money "
     "that will be spent in the UAE."),
    (60, "Amana open positions", "=B29", 15, "0",
     "Forty-one open positions on an account worth under AED 3,200 means the carrying costs on the "
     "statement — floating and overnight fees — do real damage relative to the capital at work."),
    (61, "Portfolio as a share of net worth", "=IF('Net Worth'!F28=0,0,E36/'Net Worth'!F28)",
     0.60, S.PCT,
     "How much of your net worth is actually compounding rather than sitting in a current account."),
]
for r, label, val, limit, fmt, read in risk:
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", val, fmt=fmt)
    S.put(ws, f"C{r}", limit, fmt=fmt, color=S.INPUT_BLUE)
    S.put(ws, f"D{r}", f'=IF(B{r}<=C{r},"Within limit","BREACH")', bold=True)
    S.put(ws, f"J{r}", read, wrap=True)

S.note(ws, 63,
       "The portfolio is doing its job — every scheme is in profit at the snapshot date and the SIP "
       "has been running without interruption. The two things worth fixing are structural rather "
       "than performance-related: one fund house holds nearly everything, and nearly everything is "
       "denominated in a currency you do not spend. Both are fixed with future contributions, not "
       "by selling anything.", W)
ws.row_dimensions[63].height = 52

# =============================================================== CHECKS =====
ws = wb["Checks"]
S.put(ws, "B12", "='Inputs'!B41", fmt=S.AED)
S.put(ws, "B17", "='Inputs'!B51", fmt=S.AED)
S.put(ws, "C17", 189.19, fmt=S.AED)
S.put(ws, "D17", "=B17-C17", fmt=S.AED)
S.put(ws, "E17", 0.01, fmt=S.NUM2)
S.put(ws, "F17", '=IF(ABS(D17)<=E17,"OK","CHECK")')

c = ws["A18"]
c.value = ('="Total extra cash including the September SIP: AED "&TEXT(\'Inputs\'!B53,"#,##0.00")'
           '&" — the AED "&TEXT(\'Inputs\'!B51,"#,##0.00")&" bills-and-survival gap plus AED "'
           '&TEXT(\'Inputs\'!B39,"#,##0.00")&" of SIP funding. '
           'Earlier notes in this workbook quoted AED 651.19; the correct arithmetic is AED 651.59, '
           'and this cell now computes it rather than repeating it."')
c.font = Font(name=S.FONT, sz=11, b=True, color="FFB45309")
c.alignment = Alignment(wrap_text=True, vertical="center")

HDRC = ["Check", "Actual", "Expected", "Difference", "Tolerance", "Status", "Fix / note"]
S.band(ws, 21, "EXTENDED INTEGRITY CHECKS — WEALTH, BUDGET, DEBT AND ADVISOR SHEETS", 7)
S.header(ws, 22, HDRC)
LED_D = "'Recent Ledger'!$D$6:$D$101"
LED_I = "'Recent Ledger'!$I$6:$I$101"
LED_J = "'Recent Ledger'!$J$6:$J$101"
checks = [
    ("Net Worth cash ties to Inputs liquidity", "='Net Worth'!F14", "='Inputs'!B14", 0.01,
     "Both must read the same confirmed balances."),
    ("Balance sheet identity holds", "='Net Worth'!F28", "='Net Worth'!F25-'Net Worth'!F27", 0.01,
     "Net worth must equal assets less liabilities."),
    ("Spend Analysis total ties to the ledger", "='Spend Analysis'!B17",
     f'=SUMIFS({LED_D},{LED_J},1)', 0.01,
     "Category totals must account for every flagged spend line and nothing else."),
    ("Every ledger line is classified", f'=COUNTIF({LED_I},"<>")', 96, 0,
     "All 96 ledger rows carry a budget category, including the deliberately excluded ones."),
    ("No spend line left uncategorised", f'=COUNTIFS({LED_J},1,{LED_I},"Excluded")', 0, 0,
     "A row cannot count as spending and be excluded at the same time."),
    ("Budget actuals tie to Spend Analysis",
     "=SUM(Budget!C13:C16)+SUM(Budget!C21:C26)", "='Spend Analysis'!B17", 0.01,
     "The Budget pulls the same ledger totals the Spend Analysis sheet reports."),
    ("Budget income identity holds", "=Budget!B38+Budget!B39", "=Budget!B8", 0.01,
     "Outflow plus surplus must equal income."),
    ("Allocation targets sum to 100%", "='Investments'!D50", 1, 0.0001,
     "Otherwise every drift figure on the allocation table is meaningless."),
    ("Investments AED total ties to Net Worth", "='Investments'!E36", "='Net Worth'!F21", 0.01,
     "Same holdings, same rates, two sheets."),
    ("Allocation total ties to invested total", "='Investments'!B50", "='Investments'!E36", 0.01,
     "Every dirham of the portfolio is allocated to exactly one asset class."),
    ("Debt schedule clears to zero", "='Debt Plan'!D17", 0, 0.01,
     "The recommended route must actually end the debt."),
    ("Debt payments equal the exposure", "='Debt Plan'!G17", "='Inputs'!B38", 0.01,
     "Total paid must equal total owed — no more, no less."),
    ("Wealth Plan closed form matches the table", "='Wealth Plan'!F43", "='Wealth Plan'!F38", 0.01,
     "The sensitivity block and the year-by-year projection use different maths; they must agree "
     "on the base case."),
    ("Advisor score weights sum to 100%", "='AI Advisor'!D13", 1, 0.0001,
     "Otherwise the health score is not out of 100."),
    ("Extra cash before 15 Sep computes correctly", "='Inputs'!B53", 651.59, 0.01,
     "AED 189.19 gap plus AED 462.40 SIP. Corrects the AED 651.19 quoted in earlier notes."),
    ("FX rates are reciprocal", "='FX & Assumptions'!B7*'FX & Assumptions'!B8", 1, 0.0001,
     "Guards against someone editing one rate and not the other."),
    ("Rent goal gap ties to Net Worth", "=Goals!D8", "='Net Worth'!F34", 0.01,
     "The Goals sheet and the balance sheet must agree on what is still owed on the cheque."),
    ("Ledger balance trail ends at the confirmed balance", "='Recent Ledger'!G101", "='Inputs'!B7",
     0.01,
     "The last balance in the ledger must equal the confirmed FAB 4001 balance on Inputs. This is "
     "the strongest single proof that the ledger is complete."),
]
r = 23
for label, actual, expected, tol, fix in checks:
    S.put(ws, f"A{r}", label)
    fmt = "0" if isinstance(expected, int) and not isinstance(expected, bool) and tol == 0 else S.AED
    S.put(ws, f"B{r}", actual, fmt=fmt)
    S.put(ws, f"C{r}", expected, fmt=fmt)
    S.put(ws, f"D{r}", f"=B{r}-C{r}", fmt=fmt)
    S.put(ws, f"E{r}", tol, fmt=S.NUM4)
    S.put(ws, f"F{r}", f'=IF(ABS(D{r})<=E{r},"OK","CHECK")', bold=True)
    S.put(ws, f"G{r}", fix, wrap=True)
    r += 1
LASTC = r - 1
S.put(ws, f"A{r}", "CHECKS PASSING", bold=True, fill=S.HEAD_BG)
S.put(ws, f"B{r}", f'=COUNTIF(F6:F17,"OK")+COUNTIF(F23:F{LASTC},"OK")', fmt="0",
      bold=True, fill=S.HEAD_BG)
S.put(ws, f"C{r}", f"=12+{LASTC-23+1}", fmt="0", bold=True, fill=S.HEAD_BG)
S.put(ws, f"D{r}", f"=B{r}-C{r}", fmt="0", bold=True, fill=S.HEAD_BG)
S.put(ws, f"E{r}", 0, fmt="0", bold=True, fill=S.HEAD_BG)
S.put(ws, f"F{r}", f'=IF(B{r}=C{r},"ALL OK","REVIEW")', bold=True, fill=S.HEAD_BG)
S.put(ws, f"G{r}", "If this reads REVIEW, find the CHECK rows above before trusting any number in "
                   "this workbook.", wrap=True, bold=True, fill=S.HEAD_BG)
CHECK_SUMMARY_ROW = r

# ============================================================ DASHBOARD =====
ws = wb["Dashboard"]
W = 14
S.band(ws, 29, "WEALTH POSITION", W)
S.header(ws, 30, ["Metric", "Value", "Metric", "Value", "Metric", "Value", "Metric", "Value",
                  "", "Metric", "Value", "", "", ""])
pairs = [
    (31, [("Net worth", "='Net Worth'!F28", S.AED),
          ("Total assets", "='Net Worth'!F25", S.AED),
          ("Total liabilities", "='Net Worth'!F27", S.AED),
          ("Investments (AED)", "='Net Worth'!F21", S.AED)],
     [("Health score", "='AI Advisor'!C13", S.NUM2)]),
    (32, [("Free net worth after commitments", "='Net Worth'!F36", S.AED),
          ("Free cash today", "='Inputs'!B43", S.AED),
          ("Emergency cover (months)", "='Net Worth'!B46", S.NUM2),
          ("Debt to assets", "='Net Worth'!B42", S.PCT)],
     [("Grade", "='AI Advisor'!E13", None)]),
    (33, [("Monthly plan surplus", "=Budget!B39", S.AED),
          ("Savings rate achieved", "=Budget!B42", S.PCT),
          ("Lifestyle run rate", "=Budget!D27", S.AED),
          ("Daily burn", "='Spend Analysis'!B39", S.AED)],
     [("Checks passing", f"=Checks!B{CHECK_SUMMARY_ROW}", "0")]),
    (34, [("Capital at 20-year horizon", "='Wealth Plan'!F38", S.AED),
          ("In today's money", "='Wealth Plan'!G38", S.AED),
          ("Independence target", "='Wealth Plan'!B54", S.AED),
          ("Years to independence", "='Wealth Plan'!B55", None)],
     [("Checks expected", f"=Checks!C{CHECK_SUMMARY_ROW}", "0")]),
]
for r, quad, extra in pairs:
    cols = ["A", "C", "E", "G"]
    vals = ["B", "D", "F", "H"]
    for (label, formula, fmt), lc, vc in zip(quad, cols, vals):
        S.put(ws, f"{lc}{r}", label)
        S.put(ws, f"{vc}{r}", formula, fmt=fmt)
    for label, formula, fmt in extra:
        S.put(ws, f"J{r}", label)
        S.put(ws, f"K{r}", formula, fmt=fmt)

S.band(ws, 36, "WHAT TO DO NEXT — TOP ACTIONS FROM THE AI ADVISOR", W)
ws.merge_cells(start_row=37, start_column=1, end_row=37, end_column=W)
S.put(ws, "A37", "='AI Advisor'!A47", wrap=True, fill=S.HEAD_BG, bold=True)
ws.row_dimensions[37].height = 92
S.note(ws, 39,
       "This dashboard now mirrors two things at once: the ninety-day survival plan above, and the "
       "long-term wealth position below. Read the top half when deciding what to do this week, and "
       "the bottom half when deciding whether the week mattered.", W)
ws.row_dimensions[39].height = 40
ws["A3"].value = ("Controlled mirror · Expected salary is never treated as banked cash · "
                  "Budget, Net Worth, Wealth Plan, Goals and the AI Advisor all recalculate from "
                  "the Inputs and Recent Ledger sheets")

# ==================================================== ORDER AND TAB COLOURS ==
order = ["Dashboard", "AI Advisor", "Inputs", "FX & Assumptions", "Budget", "Spend Analysis",
         "Recent Ledger", "Monthly Plan", "Rent Closure", "Obligations", "Debt Plan",
         "Investments", "Net Worth", "Wealth Plan", "Goals", "Checks", "Sources & Audit"]
assert sorted(order) == sorted(wb.sheetnames), set(order) ^ set(wb.sheetnames)
wb._sheets = [wb[n] for n in order]
TABS = {
    "Dashboard": "FF071827", "AI Advisor": "FF0D3452",
    "Inputs": "FF1F6FEB", "FX & Assumptions": "FF1F6FEB",
    "Budget": "FF0E7C66", "Spend Analysis": "FF0E7C66", "Recent Ledger": "FF0E7C66",
    "Monthly Plan": "FFB45309", "Rent Closure": "FFB45309", "Obligations": "FFB45309",
    "Debt Plan": "FFB42318",
    "Investments": "FF6941C6", "Net Worth": "FF6941C6", "Wealth Plan": "FF6941C6",
    "Goals": "FF6941C6",
    "Checks": "FF687385", "Sources & Audit": "FF687385",
}
for name, colour in TABS.items():
    wb[name].sheet_properties.tabColor = colour
wb.active = 0

wb.save(F)
print("step E ok; checks summary row", CHECK_SUMMARY_ROW)
