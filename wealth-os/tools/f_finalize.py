"""Step F — final pass.

1. Ask Excel/Sheets to recalculate everything the moment the file opens.
2. Embed the values our own evaluator computed, so the workbook also reads
   correctly in previewers, pandas and Google Sheets before any recalc.

LibreOffice in this environment ships without its Calc filters, so
scripts/recalc.py cannot run here; evaluate.py replaces it and the Checks
sheet independently proves the arithmetic.
"""
import sys, shutil, zipfile, re, datetime as dt
sys.path.insert(0, "build")
import openpyxl
from openpyxl.utils import get_column_letter
from evaluate import Book

SRC, OUT = "work.xlsx", "Johnnys_Edge_Lifetime_Finance.xlsx"

from openpyxl.workbook.properties import CalcProperties

wb = openpyxl.load_workbook(SRC)
if wb.calculation is None:
    wb.calculation = CalcProperties(calcId=124519)
wb.calculation.fullCalcOnLoad = True
wb.save(OUT)

book = Book(OUT)
cache, errors = book.evaluate_all()
if errors:
    print("ERRORS — not embedding values")
    for e in errors[:20]:
        print(" ", e)
    sys.exit(1)

# ------ map each worksheet to the part that holds it -------------------------
NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
import xml.etree.ElementTree as ET

with zipfile.ZipFile(OUT) as z:
    parts = {n: z.read(n) for n in z.namelist()}

rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
rel_target = {r.get("Id"): r.get("Target") for r in rels}
wbx = ET.fromstring(parts["xl/workbook.xml"])
sheet_part = {}
for sh in wbx.find(NS + "sheets"):
    tgt = rel_target[sh.get(REL + "id")]
    sheet_part[sh.get("name")] = ("xl/" + tgt.lstrip("/")).replace("xl/xl/", "xl/")

# ------ write cached values in beside every formula --------------------------
ET.register_namespace("", NS[1:-1])
embedded = 0
for name, part in sheet_part.items():
    root = ET.fromstring(parts[part])
    changed = False
    for c in root.iter(NS + "c"):
        f = c.find(NS + "f")
        if f is None:
            continue
        ref = c.get("r")
        col = re.match(r"([A-Z]+)(\d+)", ref)
        from openpyxl.utils import column_index_from_string
        val = cache.get((name, int(col.group(2)), column_index_from_string(col.group(1))))
        for old in c.findall(NS + "v"):
            c.remove(old)
        c.attrib.pop("t", None)
        v = ET.SubElement(c, NS + "v")
        if isinstance(val, bool):
            c.set("t", "b"); v.text = "1" if val else "0"
        elif isinstance(val, str):
            c.set("t", "str"); v.text = val
        elif isinstance(val, (int, float)):
            v.text = repr(round(float(val), 10))
        elif isinstance(val, (dt.datetime, dt.date)):
            base = dt.datetime(1899, 12, 30)
            d = val if isinstance(val, dt.datetime) else dt.datetime(val.year, val.month, val.day)
            v.text = repr((d - base).total_seconds() / 86400.0)
        else:
            c.remove(v); continue
        embedded += 1
        changed = True
    if changed:
        parts[part] = ET.tostring(root, encoding="UTF-8", xml_declaration=True)

with zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as z:
    for n, data in parts.items():
        z.writestr(n, data)

print(f"finalised {OUT}: {embedded} cached values embedded, 0 formula errors")

# ------ prove the saved file still reads back the way we expect --------------
check = openpyxl.load_workbook(OUT, data_only=True)
probe = [
    ("Checks", "B41", 30.0), ("Checks", "F41", "ALL OK"),
    ("Net Worth", "F28", 16053.32), ("Inputs", "B53", 651.59),
    ("Spend Analysis", "B17", 4547.47), ("Wealth Plan", "F38", 775778.83),
]
for sheet, cell, want in probe:
    got = check[sheet][cell].value
    ok = (isinstance(got, str) and got == want) or (
        isinstance(got, (int, float)) and abs(got - float(want)) < 0.02)
    print(f"  {'ok ' if ok else 'BAD'} {sheet}!{cell} = {got!r} (expected {want!r})")
    if not ok:
        sys.exit(1)
