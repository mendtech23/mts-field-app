"""Shared house style for Johnny's Edge Lifetime Finance workbook.

Conventions reverse-engineered from the original file so every new sheet
matches the existing ones exactly:
  title       22pt bold white on #071827, merged across the sheet width
  subtitle    10pt #687385 on #EAF0F5, wrapped
  band        11pt bold white on #0D3452, merged across the sheet width
  header      11pt bold #101827 on #EAF0F5, wrapped
  body        11pt Carlito, no fill, no borders
Colour of the *text* carries the financial-model meaning:
  blue  #0000FF  hardcoded input / lever the owner edits
  black          formula computed on this sheet
  green #008000  link to another sheet in this workbook
"""
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FONT = "Carlito"

INK        = "FF101827"
MUTED      = "FF687385"
TITLE_BG   = "FF071827"
BAND_BG    = "FF0D3452"
HEAD_BG    = "FFEAF0F5"
INPUT_BLUE = "FF0000FF"
LINK_GREEN = "FF008000"
ALERT_RED  = "FFB42318"
GOOD_GREEN = "FF067647"
LEVER_FILL = "FFFFF7C0"   # pale yellow: cells the owner is meant to edit
OK_FILL    = "FFE8F5EE"
WARN_FILL  = "FFFDECEA"

# number formats copied verbatim from the original workbook
AED   = '"AED "#,##0.00;[Red]("AED "#,##0.00);-'
AED0  = '"AED "#,##0;[Red]("AED "#,##0);-'
INR   = '"INR "#,##0.00;[Red]("INR "#,##0.00);-'
USD   = '"US$"#,##0.00;[Red]-"US$"#,##0.00'
PCT   = '0.0%;[Red](0.0%);-'
PCT2  = '0.00%;[Red](0.00%);-'
NUM4  = '#,##0.0000'
NUM2  = '#,##0.00'
DATE  = 'dd-mmm-yyyy'
DATET = 'dd mmm yyyy hh:mm'
MULT  = '0.00"x"'


def title(ws, text, width, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, sz=22, b=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=TITLE_BG)
    c.alignment = Alignment(vertical="center")
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=TITLE_BG)
    ws.row_dimensions[row].height = 29
    ws.row_dimensions[row + 1].height = 29


def subtitle(ws, text, width, row=3):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, sz=10, color=MUTED)
    c.fill = PatternFill("solid", fgColor=HEAD_BG)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=HEAD_BG)
    ws.row_dimensions[row].height = 30


def band(ws, row, text, width):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, sz=11, b=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=BAND_BG)
    c.alignment = Alignment(vertical="center")
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BAND_BG)
    ws.row_dimensions[row].height = 23


def header(ws, row, labels, col0=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=col0 + i, value=label)
        c.font = Font(name=FONT, sz=11, b=True, color=INK)
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 23


def put(ws, coord, value, fmt=None, bold=False, color=None, fill=None,
        wrap=False, size=11, italic=False):
    c = ws[coord]
    c.value = value
    c.font = Font(name=FONT, sz=size, b=bold, i=italic,
                  color=color if color else (INK if bold else None))
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="center")
    return c


def row_values(ws, row, values, col0=1, fmts=None, bold=False, colors=None,
               wrap_last=True, fill=None):
    """Write one table row. `values` may hold None to skip a cell."""
    n = len(values)
    for i, v in enumerate(values):
        if v is None:
            continue
        coord = f"{get_column_letter(col0 + i)}{row}"
        put(ws, coord, v,
            fmt=(fmts[i] if fmts else None),
            bold=bold,
            color=(colors[i] if colors else None),
            fill=fill,
            wrap=(wrap_last and i == n - 1))


def note(ws, row, text, width):
    """Full-width wrapped footnote block, styled like the originals."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, sz=10, color=MUTED)
    c.fill = PatternFill("solid", fgColor=HEAD_BG)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=HEAD_BG)


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def setup(ws):
    ws.sheet_view.showGridLines = False
