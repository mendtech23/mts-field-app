"""Step C — Net Worth and Wealth Plan sheets."""
import sys, datetime as dt
sys.path.insert(0, "build")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import style as S

F = "work.xlsx"
wb = openpyxl.load_workbook(F)

# =========================================================== NET WORTH ======
ws = wb.create_sheet("Net Worth")
S.setup(ws)
S.widths(ws, {"A": 34, "B": 24, "C": 16, "D": 11, "E": 14, "F": 18, "G": 50})
S.title(ws, "Net Worth — Everything You Own, Everything You Owe", 7)
S.subtitle(ws, "One number, in dirhams, built from confirmed balances only. Indian rupee and US "
               "dollar holdings are converted at the rates on the FX & Assumptions sheet. "
               "Expected salary is never counted as an asset.", 7)

def asset(r, name, held, native, ccy, rate, note):
    S.put(ws, f"A{r}", name)
    S.put(ws, f"B{r}", held)
    S.put(ws, f"C{r}", native, fmt={"AED": S.AED, "INR": S.INR, "USD": S.USD}[ccy],
          color=S.LINK_GREEN)
    S.put(ws, f"D{r}", ccy)
    S.put(ws, f"E{r}", rate, fmt=S.NUM4, color=S.LINK_GREEN if isinstance(rate, str) else None)
    S.put(ws, f"F{r}", f"=C{r}*E{r}", fmt=S.AED)
    S.put(ws, f"G{r}", note, wrap=True)

def total(r, name, formula, note, basis=""):
    S.put(ws, f"A{r}", name, bold=True, fill=S.HEAD_BG)
    S.put(ws, f"B{r}", basis, bold=True, fill=S.HEAD_BG)
    for col in "CDE":
        S.put(ws, f"{col}{r}", None, fill=S.HEAD_BG)
    S.put(ws, f"F{r}", formula, fmt=S.AED, bold=True, fill=S.HEAD_BG)
    S.put(ws, f"G{r}", note, wrap=True, bold=True, fill=S.HEAD_BG)

S.band(ws, 5, "ASSETS — LIQUID CASH", 7)
S.header(ws, 6, ["Asset", "Held at", "Native amount", "Currency", "Rate to AED",
                 "Value (AED)", "Note"])
asset(7, "FAB 4001 — spending account", "First Abu Dhabi Bank", "='Inputs'!B7", "AED", 1,
      "The day-to-day card account. Confirmed 25 Aug after the ENOC purchase.")
asset(8, "FAB 4002 — rent account", "First Abu Dhabi Bank", "='Inputs'!B8", "AED", 1,
      "Holds AED 5,990.45 of protected rent plus AED 100.25 towards DEWA. Spendable balance is "
      "the AED 100.25, not the AED 6,090.70.")
asset(9, "FAB emergency fund", "First Abu Dhabi Bank", "='Inputs'!B9", "AED", 1,
      "Ring-fenced. Counts as an asset but must never be counted as available cash.")
asset(10, "NBD current", "Emirates NBD", "='Inputs'!B10", "AED", 1,
      "Card-linked; effectively empty.")
asset(11, "NBD Plus Saver", "Emirates NBD", "='Inputs'!B11", "AED", 1,
      "Last reported 14 Aug. Refresh if it has moved.")
asset(12, "Tabby Cash wallet", "Tabby", "='Inputs'!B12", "AED", 1,
      "The wallet, not the Card. Spending it never touches Card debt.")
asset(13, "Cash on hand / Wio", "Cash", "='Inputs'!B13", "AED", 1,
      "No other cash balance reported.")
total(14, "TOTAL LIQUID CASH", "=SUM(F7:F13)",
      "Ties to the confirmed liquidity total on Inputs. The Checks sheet proves it.",
      "Sum of confirmed balances")

S.band(ws, 16, "ASSETS — INVESTMENTS", 7)
S.header(ws, 17, ["Asset", "Held at", "Native amount", "Currency", "Rate to AED",
                  "Value (AED)", "Note"])
asset(18, "Indian mutual funds — 6 schemes", "Nippon India / Motilal Oswal",
       "='Investments'!E12", "INR", "='FX & Assumptions'!B7",
       "Snapshot value from MF Central, 10 Aug 2026. Five Nippon schemes plus one Motilal Oswal "
       "midcap fund.")
asset(19, "ICICI settlement cash", "ICICI", "='Inputs'!B40", "INR", "='FX & Assumptions'!B7",
       "Idle rupee cash left after the August SIPs. Earning nothing; sweep it into the next SIP.")
asset(20, "Amana trading account", "Amana Capital", "='Investments'!B22", "USD",
       "='FX & Assumptions'!B9",
       "Ending account value on the statement through 10 Aug 2026, across 41 open positions.")
total(21, "TOTAL INVESTMENTS", "=SUM(F18:F20)",
      "Everything that is capable of compounding. This is the number the Wealth Plan grows.",
      "Sum of converted values")

S.band(ws, 23, "BALANCE SHEET", 7)
S.header(ws, 24, ["Item", "Basis", "Native amount", "Currency", "Rate to AED",
                  "Value (AED)", "Note"])
total(25, "TOTAL ASSETS", "=F14+F21", "Liquid cash plus investments.", "Cash + investments")
asset(26, "Tabby outstanding exposure", "Tabby Card", "='Inputs'!B38", "AED", 1,
      "AED 1,972.03 August statement plus AED 715.33 September statement. The card is frozen, so "
      "this balance can only fall from here.")
total(27, "TOTAL LIABILITIES", "=F26",
      "The only interest-bearing-style obligation in the file. Everything else is a bill, not a debt.",
      "Sum of debts")
S.put(ws, "A28", "NET WORTH", bold=True, fill=S.BAND_BG, color="FFFFFFFF")
S.put(ws, "B28", "Assets less liabilities", bold=True, fill=S.BAND_BG, color="FFFFFFFF")
for col in "CDE":
    S.put(ws, f"{col}28", None, fill=S.BAND_BG)
S.put(ws, "F28", "=F25-F27", fmt=S.AED, bold=True, fill=S.BAND_BG, color="FFFFFFFF")
S.put(ws, "G28", "The single number to watch month over month. It should rise even in a month "
                 "where the bank balance falls, because units bought are worth more than cash spent.",
      wrap=True, bold=True, fill=S.BAND_BG, color="FFFFFFFF")
ws.row_dimensions[28].height = 32

S.band(ws, 30, "COMMITTED OBLIGATIONS BEFORE 22 OCTOBER", 7)
S.header(ws, 31, ["Item", "Basis", "Native amount", "Currency", "Rate to AED",
                  "Value (AED)", "Note"])
asset(32, "October rent cheque", "Lease schedule", "='Inputs'!B30", "AED", 1,
      "Clears 22 Oct. The 26 Oct salary is four days too late to fund it.")
asset(33, "Rent already protected", "FAB 4002", "='Inputs'!B31", "AED", 1,
      "Already banked and ring-fenced.")
total(34, "RENT STILL TO FUND", "=MAX(0,F32-F33)",
      "What has to be found before 21 Oct. Not a liability on the balance sheet, but a claim on "
      "every dirham earned between now and then.", "Cheque less protected")
asset(35, "Extra cash required before 15 Sep", "Inputs funding build-up", "='Inputs'!B53", "AED", 1,
      "Bills-and-survival gap plus the September SIP.")
total(36, "FREE NET WORTH AFTER COMMITMENTS", "=F28-F34-F35",
      "Net worth with the next two months' hard commitments already subtracted. This is the "
      "honest picture of how much room actually exists.", "Net worth less commitments")

S.band(ws, 38, "COMPOSITION AND SOLVENCY", 7)
S.header(ws, 39, ["Metric", "Value", "Benchmark", "Status", "", "", "Read"])
ratios = [
    (40, "Cash as a share of assets", "=IF(F25=0,0,F14/F25)", 0.30, S.PCT,
     '=IF(B40<=C40,"OK","CASH HEAVY")',
     "Most of the cash here is spoken for by rent, so the real free-cash share is far lower than "
     "this figure suggests."),
    (41, "Investments as a share of assets", "=IF(F25=0,0,F21/F25)", 0.60, S.PCT,
     '=IF(B41>=C41,"OK","LOW")',
     "The share of the balance sheet that compounds."),
    (42, "Debt to assets", "=IF(F25=0,0,F27/F25)", 0.20, S.PCT,
     '=IF(B42<=C42,"OK","HIGH")',
     "Tabby against total assets. Manageable in size, but it is the most expensive money in the "
     "file if a fee is ever triggered."),
    (43, "Free cash outside rent and emergency", "='Inputs'!B43", 1000, S.AED,
     '=IF(B43>=C43,"OK","THIN")',
     "The genuinely spendable balance today. Everything else is protected."),
    (44, "Emergency fund held", "='Inputs'!B32", "=B45", S.AED,
     '=IF(B44>=B45,"FUNDED","UNFUNDED")',
     "Against the six-month target below. This is the largest single gap on the balance sheet."),
    (45, "Emergency fund target", "=Budget!B17*'FX & Assumptions'!B33", None, S.AED, None,
     "Six months of essential spending, taken from the Budget essentials subtotal."),
    (46, "Emergency cover held", "=IF(Budget!B17=0,0,'Inputs'!B32/Budget!B17)", 6, S.NUM2,
     '=IF(B46>=C46,"OK","CRITICAL")',
     "Months of essential spending the emergency fund would actually cover. Units are months."),
    (47, "Liquidity ratio (free cash to monthly essentials)",
     "=IF(Budget!B17=0,0,'Inputs'!B43/Budget!B17)", 1, S.NUM2,
     '=IF(B47>=C47,"OK","CRITICAL")',
     "How much of one month of essentials today's free cash would cover. Units are months."),
]
for r, label, val, bench, fmt, verdict, read in ratios:
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", val, fmt=fmt)
    if bench is not None:
        S.put(ws, f"C{r}", bench, fmt=fmt,
              color=S.INPUT_BLUE if isinstance(bench, (int, float)) else None)
    if verdict:
        S.put(ws, f"D{r}", verdict, bold=True)
    S.put(ws, f"G{r}", read, wrap=True)

S.note(ws, 49,
       "Net worth is the scoreboard; the Budget is the game. A month where spending was disciplined "
       "and the SIP was funded shows up here as a higher number even if the current account looks "
       "empty. Update the confirmed balances on Inputs, and this sheet updates itself.", 7)
ws.row_dimensions[49].height = 46

# ========================================================== WEALTH PLAN =====
ws = wb.create_sheet("Wealth Plan")
S.setup(ws)
S.widths(ws, {"A": 32, "B": 16, "C": 18, "D": 18, "E": 18, "F": 19, "G": 20, "H": 20, "I": 34})
S.title(ws, "Wealth Plan — Turning a Salary into Capital", 9)
S.subtitle(ws, "A projection, not a promise. Contributions are added at each year end and grown at "
               "the blended planning return from the FX & Assumptions sheet. Change one lever there "
               "and the whole table moves. Today's-money column strips out inflation so the numbers "
               "mean something.", 9)

S.band(ws, 5, "STARTING POSITION AND ENGINE", 9)
S.header(ws, 6, ["Input", "Value", "Source", "", "", "", "", "", "Note"])
starts = [
    (7, "Investable assets today", "='Net Worth'!F21", "Net Worth sheet", S.AED,
     "Mutual funds, broker cash and the Amana account. Bank balances are excluded: they are "
     "working capital, not capital."),
    (8, "Monthly SIP today", "=Budget!B32", "Budget sheet", S.AED,
     "The recurring investment that already exists."),
    (9, "Extra monthly investment", 0.0, "LEVER — set this yourself", S.AED,
     "Anything you redirect on top of the SIP once the plan balances. Set it to 300 and watch the "
     "horizon value below; that is what one restaurant week a month is worth over twenty years."),
    (10, "Total monthly contribution", "=B8+B9", "Formula", S.AED,
     "What actually goes in every month."),
    (11, "Year-one contribution", "=B10*12", "Formula", S.AED,
     "First full plan year."),
    (12, "Annual step-up", "='FX & Assumptions'!B32", "FX & Assumptions", S.PCT,
     "Raising the contribution each year is what separates a savings account from a wealth plan."),
    (13, "Blended planning return", "='FX & Assumptions'!B20", "FX & Assumptions", S.PCT,
     "Weighted by the current investment mix, plus the scenario adjustment."),
    (14, "Inflation", "='FX & Assumptions'!B18", "FX & Assumptions", S.PCT,
     "Used for the today's-money column."),
    (15, "Horizon", "='FX & Assumptions'!B42", "FX & Assumptions", "0",
     "Years projected below."),
]
for r, label, val, src, fmt, note in starts:
    lever_row = (r == 9)
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", val, fmt=fmt,
          color=S.INPUT_BLUE if lever_row else None,
          fill=S.LEVER_FILL if lever_row else None, bold=lever_row)
    S.put(ws, f"C{r}", src)
    S.put(ws, f"I{r}", note, wrap=True)

S.band(ws, 17, "YEAR-BY-YEAR PROJECTION", 9)
S.header(ws, 18, ["Plan year", "Calendar year", "Opening capital", "Contributions",
                  "Investment growth", "Closing capital", "Closing in today's money",
                  "Cumulative contributions", "Milestone"])
FIRST, N = 19, 20
for i in range(N):
    r = FIRST + i
    y = i + 1
    S.put(ws, f"A{r}", y, fmt="0")
    S.put(ws, f"B{r}", f"=YEAR('FX & Assumptions'!$B$41)+A{r}-1", fmt="0")
    S.put(ws, f"C{r}", "=$B$7" if i == 0 else f"=F{r-1}", fmt=S.AED)
    S.put(ws, f"D{r}", "=$B$11" if i == 0 else f"=D{r-1}*(1+$B$12)", fmt=S.AED)
    S.put(ws, f"E{r}", f"=C{r}*$B$13", fmt=S.AED)
    S.put(ws, f"F{r}", f"=C{r}+D{r}+E{r}", fmt=S.AED)
    S.put(ws, f"G{r}", f"=F{r}/(1+$B$14)^A{r}", fmt=S.AED)
    S.put(ws, f"H{r}", f"=D{r}" if i == 0 else f"=H{r-1}+D{r}", fmt=S.AED)
    S.put(ws, f"I{r}",
          f'=IF(AND(F{r}>=1000000,C{r}<1000000),"AED 1,000,000 crossed",'
          f'IF(AND(F{r}>=500000,C{r}<500000),"AED 500,000 crossed",'
          f'IF(AND(F{r}>=250000,C{r}<250000),"AED 250,000 crossed",'
          f'IF(AND(F{r}>=100000,C{r}<100000),"AED 100,000 crossed",'
          f'IF(AND(F{r}>=50000,C{r}<50000),"AED 50,000 crossed","")))))')
LAST = FIRST + N - 1
S.put(ws, f"A{LAST+1}", "HORIZON TOTAL", bold=True, fill=S.HEAD_BG)
for col in "BC":
    S.put(ws, f"{col}{LAST+1}", None, fill=S.HEAD_BG)
S.put(ws, f"D{LAST+1}", f"=SUM(D{FIRST}:D{LAST})", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"E{LAST+1}", f"=SUM(E{FIRST}:E{LAST})", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"F{LAST+1}", f"=F{LAST}", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"G{LAST+1}", f"=G{LAST}", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"H{LAST+1}", f"=H{LAST}", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"I{LAST+1}", "Growth above contributions is the part the market pays you. If it is "
                        "smaller than contributions, time is the missing ingredient, not returns.",
      wrap=True, bold=True, fill=S.HEAD_BG)

SB = LAST + 3          # sensitivity band row
S.band(ws, SB, "WHAT EACH LEVER IS ACTUALLY WORTH", 9)
S.header(ws, SB + 1, ["Scenario", "Monthly contribution", "Step-up", "Return", "Horizon",
                      "Capital at horizon", "In today's money", "Difference vs base", "Read"])

def fv(p0, c1, r_, g_, n_):
    """Closed-form future value of a growing annuity with year-end payments.
    Matches the table above exactly, including the r = g edge case."""
    return (f"={p0}*(1+{r_})^{n_}+IF(ABS({r_}-{g_})<0.0001,"
            f"{c1}*{n_}*(1+{r_})^({n_}-1),"
            f"{c1}*((1+{r_})^{n_}-(1+{g_})^{n_})/({r_}-{g_}))")

scen = [
    ("Base — today's SIP, 10% step-up", "=$B$10", "=$B$12", "=$B$13",
     "Exactly the table above."),
    ("Add AED 300 a month", "=$B$10+300", "=$B$12", "=$B$13",
     "Roughly one week of restaurant spending redirected. Compare the difference column with the "
     "sacrifice; this is the cheapest wealth in the file."),
    ("Step-up 15% instead of 10%", "=$B$10", "=0.15", "=$B$13",
     "Costs nothing today. Every raise partly goes to the SIP instead of to lifestyle."),
    ("No step-up at all", "=$B$10", "=0", "=$B$13",
     "The same contribution forever. Inflation quietly shrinks it every year."),
    ("Returns 3 points lower", "=$B$10", "=$B$12", "=$B$13-0.03",
     "The stress case. Note how much less it costs you than skipping contributions does — the "
     "contribution lever is under your control, the return lever is not."),
    ("Pause the SIP for one year, then resume", "=$B$10", "=$B$12", "=$B$13",
     "Modelled as one year less of compounding on the whole horizon. A pause is never free."),
]
r = SB + 2
BASE = r
for i, (name, c, g, ret, read) in enumerate(scen):
    n_expr = "$B$15" if i != 5 else "($B$15-1)"
    S.put(ws, f"A{r}", name)
    S.put(ws, f"B{r}", c, fmt=S.AED)
    S.put(ws, f"C{r}", g, fmt=S.PCT, color=S.INPUT_BLUE if g.startswith("=0") else None)
    S.put(ws, f"D{r}", ret, fmt=S.PCT)
    S.put(ws, f"E{r}", f"={n_expr}", fmt="0")
    S.put(ws, f"F{r}", fv("$B$7", f"B{r}*12", f"D{r}", f"C{r}", f"E{r}"), fmt=S.AED)
    S.put(ws, f"G{r}", f"=F{r}/(1+$B$14)^E{r}", fmt=S.AED)
    S.put(ws, f"H{r}", f"=F{r}-F${BASE}", fmt=S.AED)
    S.put(ws, f"I{r}", read, wrap=True)
    r += 1

FIB = r + 1
S.band(ws, FIB, "FINANCIAL INDEPENDENCE", 9)
S.header(ws, FIB + 1, ["Metric", "Value", "Basis", "", "", "", "", "", "Read"])
fi = [
    ("Annual essential spending today", "=Budget!B17*12", "Budget essentials x 12", S.AED,
     "Rent accrual, utilities, groceries, fuel and family support. Not lifestyle."),
    ("Safe withdrawal rate", 0.04, "LEVER — standard planning rule", S.PCT,
     "The share of capital you could draw each year without running it down. Conservative."),
    ("Capital needed for independence", f"=IF(B{FIB+3}=0,0,B{FIB+2}/B{FIB+3})",
     "Essential spending divided by the withdrawal rate", S.AED,
     "The number at which essential living costs are paid by capital rather than by work."),
    ("Years to reach it at the base plan",
     f"=IF(COUNTIF(G{FIRST}:G{LAST},\"<\"&B{FIB+4})>={N},\"beyond horizon\","
     f"COUNTIF(G{FIRST}:G{LAST},\"<\"&B{FIB+4})+1)",
     "First year the today's-money column clears the target", "General",
     "Counted in today's money, so the answer is honest. If it reads beyond horizon, the fix is "
     "the contribution lever, not a better fund."),
    ("Capital at horizon in today's money", f"=G{LAST}", "Year-by-year projection", S.AED,
     "Where the base plan actually lands."),
    ("Shortfall against independence", f"=MAX(0,B{FIB+4}-B{FIB+6})",
     "Target less projected", S.AED,
     "Zero means the base plan gets there within the horizon."),
]
r = FIB + 2
for label, val, basis, fmt, read in fi:
    is_lever = isinstance(val, float)
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", val, fmt=fmt,
          color=S.INPUT_BLUE if is_lever else None,
          fill=S.LEVER_FILL if is_lever else None, bold=is_lever)
    S.put(ws, f"C{r}", basis)
    S.put(ws, f"I{r}", read, wrap=True)
    r += 1

S.note(ws, r + 1,
       "The uncomfortable arithmetic: the plan below only runs if the Budget balances first. Rent "
       "accrual, then the emergency fund, then debt, then the SIP — in that order. A wealth plan "
       "funded by a frozen credit card is not a wealth plan.", 9)
ws.row_dimensions[r + 1].height = 40

wb.save(F)
print("step C ok; FI rows at", FIB + 2, "to", r - 1)
