"""A small spreadsheet evaluator for the exact formula subset this workbook
uses. LibreOffice ships without its Calc filters in this environment, so
recalc.py cannot run here; this recomputes every formula from first
principles instead, which both verifies the arithmetic and produces cached
values to embed in the delivered file.

Supported: + - * / ^ % ( ) comparisons, & concatenation, absolute and
relative refs, ranges, cross-sheet refs, and the 16 functions the workbook
actually calls.
"""
import re, math, datetime as dt
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string

EPOCH = dt.datetime(1899, 12, 30)


def to_serial(v):
    if isinstance(v, dt.datetime):
        return (v - EPOCH).total_seconds() / 86400.0
    if isinstance(v, dt.date):
        return (dt.datetime(v.year, v.month, v.day) - EPOCH).days
    return v


class XlError(Exception):
    def __init__(self, code): super().__init__(code); self.code = code


# --------------------------------------------------------------- tokenizer --
TOKEN = re.compile(r"""
    (?P<str>"(?:[^"]|"")*")
  | (?P<sheet>'(?:[^']|'')*'!|[A-Za-z_][A-Za-z0-9_.]*!)
  | (?P<ref>\$?[A-Z]{1,3}\$?[0-9]{1,7}(?::\$?[A-Z]{1,3}\$?[0-9]{1,7})?)
  | (?P<func>[A-Z][A-Z0-9.]*(?=\s*\())
  | (?P<num>\d+(?:\.\d+)?(?:[eE][+-]?\d+)?|\.\d+)
  | (?P<op><=|>=|<>|[-+*/^&<>=%])
  | (?P<lp>\()
  | (?P<rp>\))
  | (?P<comma>,)
  | (?P<ws>\s+)
""", re.X)


def tokenize(f):
    out, i = [], 0
    while i < len(f):
        m = TOKEN.match(f, i)
        if not m:
            raise XlError(f"#PARSE! at {i}: {f[i:i+20]!r}")
        i = m.end()
        kind = m.lastgroup
        if kind == "ws":
            continue
        out.append((kind, m.group()))
    return out


# ------------------------------------------------------------------ parser --
class Parser:
    """Recursive descent; precedence follows Excel."""

    def __init__(self, tokens, sheet, book):
        self.t, self.i, self.sheet, self.book = tokens, 0, sheet, book

    def peek(self): return self.t[self.i] if self.i < len(self.t) else (None, None)
    def next(self):
        tok = self.peek(); self.i += 1; return tok

    def parse(self):
        v = self.compare()
        if self.i != len(self.t):
            raise XlError(f"#PARSE! trailing {self.t[self.i:]}")
        return v

    def compare(self):
        left = self.concat()
        while self.peek()[0] == "op" and self.peek()[1] in ("=", "<>", "<", ">", "<=", ">="):
            op = self.next()[1]
            right = self.concat()
            left = compare(op, left, right)
        return left

    def concat(self):
        left = self.additive()
        while self.peek()[0] == "op" and self.peek()[1] == "&":
            self.next()
            left = as_text(left) + as_text(self.additive())
        return left

    def additive(self):
        left = self.multiplicative()
        while self.peek()[0] == "op" and self.peek()[1] in "+-":
            op = self.next()[1]
            right = self.multiplicative()
            f = (lambda x, y: num(x) + num(y)) if op == "+" else (lambda x, y: num(x) - num(y))
            left = broadcast(f, left, right)
        return left

    def multiplicative(self):
        left = self.power()
        while self.peek()[0] == "op" and self.peek()[1] in "*/":
            op = self.next()[1]
            right = self.power()
            if op == "*":
                left = broadcast(lambda x, y: num(x) * num(y), left, right)
            else:
                def _div(x, y):
                    d = num(y)
                    if d == 0: raise XlError("#DIV/0!")
                    return num(x) / d
                left = broadcast(_div, left, right)
        return left

    def power(self):
        left = self.unary()
        while self.peek()[0] == "op" and self.peek()[1] == "^":
            self.next()
            left = num(left) ** num(self.unary())
        return left

    def unary(self):
        k, v = self.peek()
        if k == "op" and v in "+-":
            self.next()
            x = self.unary()
            return -num(x) if v == "-" else num(x)
        return self.postfix()

    def postfix(self):
        v = self.primary()
        while self.peek() == ("op", "%"):
            self.next()
            v = num(v) / 100.0
        return v

    def primary(self):
        k, v = self.next()
        if k == "num":  return float(v)
        if k == "str":  return v[1:-1].replace('""', '"')
        if k == "lp":
            inner = self.compare()
            if self.next()[0] != "rp": raise XlError("#PARSE! missing )")
            return inner
        if k == "func":
            if self.next()[0] != "lp": raise XlError("#PARSE! missing (")
            args = self.arglist()
            return call(v, args, self.book)
        if k == "sheet":
            name = v[:-1]
            if name.startswith("'"): name = name[1:-1].replace("''", "'")
            k2, v2 = self.next()
            if k2 != "ref": raise XlError("#REF!")
            return self.book.resolve(name, v2)
        if k == "ref":
            return self.book.resolve(self.sheet, v)
        raise XlError(f"#PARSE! unexpected {k} {v!r}")

    def arglist(self):
        args = []
        if self.peek()[0] == "rp":
            self.next(); return args
        while True:
            # an omitted argument (",,") evaluates to blank
            if self.peek()[0] in ("comma", "rp"):
                args.append(None)
            else:
                args.append(self.compare())
            k, _ = self.next()
            if k == "rp": return args
            if k != "comma": raise XlError("#PARSE! bad argument list")


# --------------------------------------------------------------- coercion --
class Range(list):
    """A rectangular range: a flat list of values, kept distinct from a scalar."""


def num(v):
    if isinstance(v, Range):
        v = v[0] if v else 0
    v = to_serial(v)
    if v is None or v == "": return 0.0
    if isinstance(v, bool): return 1.0 if v else 0.0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        try: return float(v)
        except ValueError: raise XlError("#VALUE!")
    raise XlError("#VALUE!")


def as_text(v):
    if isinstance(v, Range): v = v[0] if v else ""
    if v is None: return ""
    if isinstance(v, bool): return "TRUE" if v else "FALSE"
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    return str(v)


def scalar(v):
    return (v[0] if v else None) if isinstance(v, Range) else v


def broadcast(fn, a, b):
    """Excel operators are elementwise when either side is a range."""
    if isinstance(a, Range) or isinstance(b, Range):
        la = a if isinstance(a, Range) else None
        lb = b if isinstance(b, Range) else None
        n = max(len(la) if la is not None else 0, len(lb) if lb is not None else 0)
        return Range([fn(la[i] if la is not None else a,
                         lb[i] if lb is not None else b) for i in range(n)])
    return fn(a, b)


def compare(op, a, b):
    if isinstance(a, Range) or isinstance(b, Range):
        return broadcast(lambda x, y: compare(op, x, y), a, b)
    a, b = scalar(a), scalar(b)
    if isinstance(a, str) or isinstance(b, str):
        if not (isinstance(a, str) and isinstance(b, str)):
            a, b = as_text(a).upper(), as_text(b).upper()
        else:
            a, b = a.upper(), b.upper()
    else:
        a, b = num(a), num(b)
    return {"=": a == b, "<>": a != b, "<": a < b, ">": a > b, "<=": a <= b, ">=": a >= b}[op]


CRIT = re.compile(r"^(<=|>=|<>|<|>|=)?(.*)$", re.S)


def matches(value, criterion):
    if isinstance(criterion, Range): criterion = scalar(criterion)
    if criterion is None: return value in (None, "")
    if isinstance(criterion, (int, float)) and not isinstance(criterion, bool):
        return num(value if value is not None else 0) == float(criterion)
    op, rest = CRIT.match(str(criterion)).groups()
    op = op or "="
    if rest == "" and op == "<>":
        return value not in (None, "")
    try:
        target = float(rest)
        return compare(op, num(value if value is not None else 0), target)
    except ValueError:
        pass
    return compare(op, as_text(value), rest)


# -------------------------------------------------------------- functions --
def flat(args):
    out = []
    for a in args:
        if isinstance(a, Range): out.extend(a)
        else: out.append(a)
    return out


def numbers(args):
    return [num(x) for x in flat(args)
            if x is not None and x != "" and not isinstance(x, str)]


def call(name, args, book):
    if name == "IF":
        cond = scalar(args[0])
        cond = bool(cond) if isinstance(cond, bool) else num(cond) != 0
        if cond: return scalar(args[1]) if len(args) > 1 else True
        return scalar(args[2]) if len(args) > 2 else False
    if name == "AND": return all(bool(x) if isinstance(x, bool) else num(x) != 0 for x in flat(args))
    if name == "OR":  return any(bool(x) if isinstance(x, bool) else num(x) != 0 for x in flat(args))
    if name == "NOT": return not (num(args[0]) != 0)
    if name == "SUM": return sum(numbers(args))
    if name == "MAX": return max(numbers(args), default=0.0)
    if name == "MIN": return min(numbers(args), default=0.0)
    if name == "ABS": return abs(num(args[0]))
    if name == "INT": return float(math.floor(num(args[0])))
    if name == "ROUND":
        d = int(num(args[1])) if len(args) > 1 else 0
        x = num(args[0]); f = 10 ** d
        return math.floor(abs(x) * f + 0.5) / f * (1 if x >= 0 else -1)
    if name == "AVERAGE":
        n = numbers(args)
        if not n: raise XlError("#DIV/0!")
        return sum(n) / len(n)
    if name == "YEAR":  return float((EPOCH + dt.timedelta(days=num(args[0]))).year)
    if name == "MONTH": return float((EPOCH + dt.timedelta(days=num(args[0]))).month)
    if name == "DAY":   return float((EPOCH + dt.timedelta(days=num(args[0]))).day)
    if name == "DATE":
        d = dt.datetime(int(num(args[0])), int(num(args[1])), int(num(args[2])))
        return (d - EPOCH).days * 1.0
    if name == "TODAY": return float((dt.datetime(2026, 8, 25) - EPOCH).days)
    if name == "IFERROR":
        try: return scalar(args[0])
        except XlError: return scalar(args[1])
    if name == "TEXT":  return excel_text(args[0], as_text(args[1]))
    if name == "COUNTIF":
        rng, crit = args[0], scalar(args[1])
        return float(sum(1 for v in (rng if isinstance(rng, Range) else [rng]) if matches(v, crit)))
    if name == "COUNTIFS":
        pairs = [(args[i], scalar(args[i + 1])) for i in range(0, len(args), 2)]
        n = len(pairs[0][0])
        return float(sum(1 for i in range(n) if all(matches(r[i], c) for r, c in pairs)))
    if name == "SUMIF":
        rng, crit = args[0], scalar(args[1])
        tot = args[2] if len(args) > 2 else rng
        return float(sum(num(tot[i]) for i in range(len(rng)) if matches(rng[i], crit)))
    if name == "SUMIFS":
        tot = args[0]
        pairs = [(args[i], scalar(args[i + 1])) for i in range(1, len(args), 2)]
        return float(sum(num(tot[i]) for i in range(len(tot))
                         if all(matches(r[i], c) for r, c in pairs)))
    if name == "SUMPRODUCT":
        arrays = [a if isinstance(a, Range) else Range([a]) for a in args]
        n = min(len(a) for a in arrays)
        return float(sum(math.prod(num(a[i]) for a in arrays) for i in range(n)))
    raise XlError(f"#NAME? {name}")


NUMFMT = re.compile(r"^#,##0(\.0+)?$")


def excel_text(v, fmt):
    x = num(v)
    m = NUMFMT.match(fmt)
    if m:
        d = len(m.group(1)) - 1 if m.group(1) else 0
        return f"{x:,.{d}f}"
    if fmt in ("0", "#"): return f"{x:,.0f}".replace(",", "")
    if fmt.endswith("%"): return f"{x * 100:.1f}%"
    return f"{x:,.2f}"


# ------------------------------------------------------------------- book --
class Book:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path)
        self.cache = {}
        self.stack = []
        self.errors = []

    def cell_value(self, sheet, row, col):
        key = (sheet, row, col)
        if key in self.cache:
            return self.cache[key]
        if key in self.stack:
            raise XlError("#CIRC!")
        raw = self.wb[sheet].cell(row=row, column=col).value
        if isinstance(raw, str) and raw.startswith("="):
            self.stack.append(key)
            try:
                val = Parser(tokenize(raw[1:]), sheet, self).parse()
                val = scalar(val)
            except XlError as e:
                val = e.code
                self.errors.append((sheet, f"{get_column_letter(col)}{row}", raw, e.code))
            except Exception as e:               # noqa: BLE001 - report, don't crash
                val = "#ERR!"
                self.errors.append((sheet, f"{get_column_letter(col)}{row}", raw, repr(e)))
            finally:
                self.stack.pop()
        else:
            val = raw
        self.cache[key] = val
        return val

    def resolve(self, sheet, ref):
        if sheet not in self.wb.sheetnames:
            raise XlError("#REF!")
        mn_c, mn_r, mx_c, mx_r = range_boundaries(ref.replace("$", ""))
        if (mn_c, mn_r) == (mx_c, mx_r):
            return self.cell_value(sheet, mn_r, mn_c)
        return Range([self.cell_value(sheet, r, c)
                      for r in range(mn_r, mx_r + 1) for c in range(mn_c, mx_c + 1)])

    def evaluate_all(self):
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        self.cell_value(ws.title, cell.row, cell.column)
        return self.cache, self.errors

    def get(self, addr):
        sheet, ref = addr.split("!")
        sheet = sheet.strip("'")
        return self.resolve(sheet, ref)
