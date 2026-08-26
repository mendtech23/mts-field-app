"""Step A — repair broken cells and complete the truncated rows in the
original workbook. Nothing here invents data: every fix is derived from a
value already present elsewhere in the file.
"""
import sys, datetime as dt
sys.path.insert(0, "build")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import style as S

SRC, OUT = "orig.xlsx", "work.xlsx"
wb = openpyxl.load_workbook(SRC)

EPOCH = dt.datetime(1899, 12, 30)
def serial(n):
    return EPOCH + dt.timedelta(days=float(n))

# ---------------------------------------------------------------- Inputs ----
ws = wb["Inputs"]

# B42 held AED 715.33 but Excel had typed the cell as a date, so it rendered
# as 15-Dec-1901 (715.33 days after the 1900 epoch). D42 held the due date as
# a raw serial. Both are the same numbers, correctly typed.
S.put(ws, "B42", 715.33, fmt=S.AED)
S.put(ws, "D42", dt.datetime(2026, 10, 3), fmt=S.DATE)
ws["F42"].value = ("September Tabby statement; confirmed amount due 3 Oct 2026. "
                   "Cell repaired: the amount had been typed as a date and displayed as 15-Dec-1901.")

# Row 44 was a merged callout with a truncated sentence. Rebuild it as a live
# sentence driven by the funding build-up added at rows 46-53 below.
c = ws["A44"]
c.value = ('="Extra cash required before 15 Sep: AED "&TEXT(B53,"#,##0.00")'
           '&" — AED "&TEXT(B51,"#,##0.00")&" to close the bills-and-survival gap plus AED "'
           '&TEXT(B52,"#,##0.00")&" to fund the 10 Sep Nippon SIP. '
           'Bank it as real cash or pause the SIP; never take it from the AED "'
           '&TEXT(B31,"#,##0.00")&" protected rent or from the emergency fund."')
c.font = Font(name=S.FONT, sz=11, b=True, color="FFFFFFFF")
c.alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[44].height = 46

S.band(ws, 46, "NEAR-TERM FUNDING GAP — CASH NEEDED BEFORE 15 SEP", 6)
S.header(ws, 47, ["Item", "Amount", "Status", "Date", "Rule", "Audit note"])

rows = [
    (48, "Survival allowance to 25 Sep", 160, "Control cap", dt.datetime(2026, 9, 25),
     "AED 5/day; AED 35/week",
     "Damage-control cap held from 26 Aug to 25 Sep. Editable lever: raising it widens the gap one-for-one."),
    (49, "Near-term bills to 15 Sep", "=B33+B36+B34+B35", "Formula", dt.datetime(2026, 9, 15),
     "Essential bills only",
     "DEWA AED 813.28 + Tabby no-fee minimum AED 1,314.50 + du AED 590.98 + Etisalat AED 323.95."),
    (50, "Cash available incl. expected payday", "=B43+B22", "Formula", dt.datetime(2026, 8, 26),
     "Only banked cash closes the gap",
     "Loose cash outside rent and emergency (AED 107.52) plus the AED 2,906 expected 26 Aug payday."),
    (51, "Bills + survival funding gap", "=MAX(0,B49+B48-B50)", "Formula", dt.datetime(2026, 9, 15),
     "Close before 15 Sep",
     "Requirement AED 3,202.71 against AED 3,013.52 available."),
    (52, "September SIP funding", "=B39", "Link", dt.datetime(2026, 9, 10),
     "Pause before borrowing",
     "Mirrors the SIP funding row above so the total below moves when the SIP estimate changes."),
    (53, "EXTRA CASH REQUIRED BEFORE 15 SEP", "=B51+B52", "Formula", dt.datetime(2026, 9, 15),
     "Bank it or pause the SIP",
     "Gap plus SIP. Note: earlier notes in this file quoted AED 651.19; the correct arithmetic is "
     "AED 189.19 + AED 462.40 = AED 651.59."),
]
for r, label, amt, status, date, rule, audit in rows:
    bold = (r == 53)
    fill = S.HEAD_BG if bold else None
    S.put(ws, f"A{r}", label, bold=bold, fill=fill)
    S.put(ws, f"B{r}", amt, fmt=S.AED, bold=bold, fill=fill,
          color=(S.INPUT_BLUE if isinstance(amt, (int, float)) else None))
    S.put(ws, f"C{r}", status, bold=bold, fill=fill)
    S.put(ws, f"D{r}", date, fmt=S.DATE, bold=bold, fill=fill)
    S.put(ws, f"E{r}", rule, bold=bold, fill=fill)
    S.put(ws, f"F{r}", audit, bold=bold, fill=fill, wrap=True)

# ---------------------------------------------------------- Rent Closure ----
ws = wb["Rent Closure"]
S.put(ws, "E20", dt.datetime(2026, 10, 21), fmt=S.DATE)   # was the raw serial 46316

# --------------------------------------------------------- Recent Ledger ----
ws = wb["Recent Ledger"]
S.put(ws, "A43", dt.datetime(2026, 8, 12, 15, 20), fmt="dd-mmm hh:mm")  # was serial 46246

# --------------------------------------------------------- Sources & Audit --
ws = wb["Sources & Audit"]
# Rows 25-32 had drifted one or two columns out of alignment with the header
# (ID | Item | Value / scope | Status | As of | Source | Notes). Rewritten in
# place with every original fact preserved and the missing IDs S17-S21 filled.
realigned = [
    (25, "S17", "12 Aug FAB 4001 card charges", "AED 49.50", "Actual",
     dt.datetime(2026, 8, 12, 21, 19), "76232DAC-B5CB-4A09-9785-FE06CB91B65C.jpeg",
     "Three FAB 4001 charges: KFC AED 33.00 household, KFC AED 1.50 household, "
     "ASAS AED 15.00 personal lifestyle; closing balance AED 150.50"),
    (26, "S18", "FAB 4001 closing balance", "AED 140.50", "Actual",
     dt.datetime(2026, 8, 13, 20, 1), "User-provided FAB SMS",
     "Nad Al Hamar Baker AED 10.00 purchase; confirmed actual; Personal Lifestyle"),
    (27, "S19", "FAB 4001 closing balance", "AED 126.50", "Actual",
     dt.datetime(2026, 8, 13, 20, 59), "User-provided FAB SMS",
     "Asas Al Madina AED 14.00 purchase; confirmed actual; Personal Lifestyle"),
    (28, "S20", "FAB 4001 balance", "AED 113.00", "Actual",
     dt.datetime(2026, 8, 14, 18, 8), "User-provided FAB SMS",
     "Emarat shop AED 13.50 purchase; confirmed actual; Personal Lifestyle"),
    (29, "S21", "FAB 4001 closing balance", "AED 63.00", "Actual",
     dt.datetime(2026, 8, 14, 18, 11), "User-provided FAB SMS",
     "Emarat fuel AED 50.00 purchase; confirmed actual; essential fuel"),
    (30, "S22", "FAB 4001 closing balance", "AED 61.00", "Actual",
     dt.datetime(2026, 8, 15, 2, 31), "User-provided FAB SMS",
     "Jackson Trading AED 2.00 purchase; confirmed actual; Personal Lifestyle"),
    (31, "S23", "FAB 4001 closing balance", "AED 58.00", "Actual",
     dt.datetime(2026, 8, 15, 2, 33), "User-provided FAB SMS",
     "Jackson Trading AED 3.00 purchase; confirmed actual; Personal Lifestyle"),
    (32, "S24", "October Tabby payment", "AED 715.33", "Actual",
     dt.datetime(2026, 10, 3), "2BBCE8E1-D085-4593-BC43-D4C6D27B2C3B.png",
     "Generated September 2026 statement; due 3 Oct 2026; actual amount; replaces AED 0 placeholder"),
]
for r, sid, item, val, status, asof, source, notes in realigned:
    S.put(ws, f"A{r}", sid)
    S.put(ws, f"B{r}", item)
    S.put(ws, f"C{r}", val)
    S.put(ws, f"D{r}", status)
    S.put(ws, f"E{r}", asof, fmt=S.DATET)
    S.put(ws, f"F{r}", source, wrap=True)
    S.put(ws, f"G{r}", notes, wrap=True)

# Rows 33-44 carried the "As of" timestamp as a raw serial number.
for r in range(33, 45):
    v = ws[f"E{r}"].value
    if isinstance(v, (int, float)):
        S.put(ws, f"E{r}", serial(v), fmt=S.DATET)

wb.save(OUT)
print("step A ok ->", OUT)
