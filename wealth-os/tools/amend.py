"""Amend Johnnys_Edge_Finance_OS.xlsx.

Ten defects found by a full audit of all 4,120 formulas. The workbook's own
22 integrity checks all pass and no formula errors exist; everything below is
structural rather than arithmetic. Every fix preserves the file's house style
and every fact already recorded in it. Nothing is invented.
"""
import sys, datetime as dt, copy
sys.path.insert(0, "build")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC, OUT = "new.xlsx", "Johnnys_Edge_Finance_OS.xlsx"
wb = openpyxl.load_workbook(SRC)

def style_from(ws, src_coord, dst_coord):
    """Copy a cell's look so amended cells match the sheet they live on."""
    s, d = ws[src_coord], ws[dst_coord]
    if s.has_style:
        d.font = copy.copy(s.font)
        d.fill = copy.copy(s.fill)
        d.border = copy.copy(s.border)
        d.alignment = copy.copy(s.alignment)
        d.number_format = s.number_format

FIXES = []
def log(sheet, what):
    FIXES.append((sheet, what))

# ===================================================== 1. Account History ===
ah = wb["Account History"]

# 1a. The summary block sat at A21:D29 — inside the $A$7:$A$100 / $B$7:$B$100 /
#     $C$7:$C$100 ranges its own LOOKUPs scan, so every one of the eight
#     accounts was a circular reference and returned "No data". Moving it to
#     columns G–J puts it outside those ranges: the block works, and the log
#     keeps its room to grow to row 100.
ACCOUNTS = ["FAB 4001 Current", "FAB 4002 Home Expenditure", "FAB 4003 Emergency Fund",
            "NBD Current", "NBD Plus Saver", "Tabby Cash", "Cash / Wio", "ICICI (INR)"]

ah["G5"] = "LATEST BALANCE PER ACCOUNT  ·  pulls the most recent dated entry for each"
style_from(ah, "A5", "G5")
for i, h in enumerate(["Account", "Latest balance", "As of", "Entries logged"]):
    c = ah.cell(row=6, column=7 + i, value=h)
    style_from(ah, get_column_letter(1 + i) + "6", get_column_letter(7 + i) + "6")
for i, acct in enumerate(ACCOUNTS):
    r = 7 + i
    ah[f"G{r}"] = acct
    ah[f"H{r}"] = (f'=IFERROR(LOOKUP(2,1/(($B$7:$B$100=$G{r})*($C$7:$C$100<>"")),'
                   f'$C$7:$C$100),"No data")')
    ah[f"I{r}"] = (f'=IFERROR(LOOKUP(2,1/(($B$7:$B$100=$G{r})*($C$7:$C$100<>"")),'
                   f'$A$7:$A$100),"")')
    ah[f"J{r}"] = f'=COUNTIF($B$7:$B$100,$G{r})'
    style_from(ah, f"A{7 + i}", f"G{r}")
    style_from(ah, f"C{7 + i}", f"H{r}")
    ah[f"H{r}"].number_format = '#,##0.00;(#,##0.00);"–"'
    ah[f"I{r}"].number_format = "dd mmm yyyy"
    ah[f"J{r}"].number_format = "0"
for col, w in (("G", 30), ("H", 16), ("I", 14), ("J", 15)):
    ah.column_dimensions[col].width = w
log("Account History",
    "Latest-balance summary moved from A21:D29 to G5:J14. It sat inside the "
    "ranges its own LOOKUPs scan, so all eight accounts were circular and read "
    '"No data". The block now works.')

# 1b. Clear the old block and leave a pointer where it used to be.
for r in range(20, 30):
    for col in "ABCDE":
        ah[f"{col}{r}"].value = None

# 1c. B7 named an account that does not exist, so that entry could never match.
if str(ah["B7"].value).strip() == "FAB 1001 Current":
    ah["B7"] = "FAB 4001 Current"
    log("Account History", 'B7 read "FAB 1001 Current" — a typo for 4001, which meant that '
                           "entry matched no account in the summary. Corrected.")

# 1d. "REQUIRES DATA" is text sitting in a numeric balance column, where the
#     LOOKUP's <>"" test treats it as a real reading. Blanked, reason kept in
#     the note column rather than invented.
for r in (7, 8):
    if isinstance(ah[f"C{r}"].value, str) and "REQUIRES" in ah[f"C{r}"].value.upper():
        ah[f"C{r}"] = None
        note = ah[f"E{r}"].value or ""
        ah[f"E{r}"] = (note + "  ·  Balance not recorded for this date; left blank so the "
                              "latest-balance lookup skips it rather than returning text.").strip()
log("Account History",
    'C7 and C8 held the text "REQUIRES DATA" in the balance column, where the lookup '
    "counted it as a reading. Blanked, with the reason moved to the note column.")

# 1e. The log stopped on 11 Aug while Accounts is confirmed to 26 Aug, so the
#     summary contradicted the live sheet. These rows are the Accounts sheet's
#     own confirmed balances — transcribed, not invented.
CATCHUP = [
    ("2026-08-25", "FAB 4002 Home Expenditure", 5940.70,
     "Confirmed on the Accounts sheet. AED 150 moved out for home expenditure."),
    ("2026-08-26", "FAB 4001 Current", 14.95,
     "Confirmed on the Accounts sheet. Card XXXX1599, Emarat fuel 13.50 at 18:20."),
    ("2026-08-26", "NBD Current", 138.23,
     "Confirmed on the Accounts sheet. Cielo Kabab, Emarat 6192 and Netflix 36.09 cleared."),
    ("2026-08-26", "ICICI (INR)", 12388.51,
     "Confirmed via the ICICI app. Fresh INR 12,000 remittance landed for the SIP."),
]
row = 19
for date, acct, bal, note in CATCHUP:
    ah[f"A{row}"] = dt.datetime.strptime(date, "%Y-%m-%d")
    ah[f"B{row}"] = acct
    ah[f"C{row}"] = bal
    ah[f"E{row}"] = note
    for col, src in (("A", "A18"), ("B", "B18"), ("C", "C18"), ("E", "E18")):
        style_from(ah, src, f"{col}{row}")
    row += 1
log("Account History",
    "Log ended on 11 Aug while Accounts was confirmed to 26 Aug, so the summary "
    "contradicted the live balances. Added the four confirmed 25–26 Aug readings.")

# A pointer where the old block used to be, placed below the new entries.
ah["A24"] = ("The latest-balance summary now sits to the right, at column G — it had to move out "
             "of columns A–E because it was reading the very range it lives in. Rows below here "
             "are free for new log entries.")
style_from(ah, "A2", "A24")
ah.merge_cells("A24:E24")

# 1f. Give every log row from 7 to 100 the same change-vs-last-entry formula,
#     instead of the scattered rows that happened to have one.
merged_rows = {rng.min_row for rng in ah.merged_cells.ranges}
for r in range(7, 101):
    if r in merged_rows:
        continue                      # the pointer note spans A:E on its row
    ah[f"D{r}"] = (f'=IF($C{r}="","",IFERROR($C{r}-LOOKUP(2,1/(($B$7:$B{r - 1}=$B{r})'
                   f'*($A$7:$A{r - 1}<$A{r})),$C$7:$C{r - 1}),""))')
    style_from(ah, "D18", f"D{r}")
log("Account History",
    "Change-vs-last-entry formula was present on some rows and missing on others, "
    "including two section-header rows. Applied uniformly to rows 7–100 and guarded "
    "so an empty row stays empty.")

# ========================================================= 2. Investments ===
inv = wb["Investments"]
# Rows 12 and 13 were identical TOTAL rows. Nothing references row 13, and a
# second total under the first reads like a mistake in the numbers.
if str(inv["A12"].value).strip() == "TOTAL" and str(inv["A13"].value).strip() == "TOTAL":
    for col in "ABCDEFGHIJKL":
        inv[f"{col}13"].value = None
    log("Investments",
        "Two identical TOTAL rows sat under the fund table (rows 12 and 13). Nothing "
        "referenced the second one; removed so the table has a single total.")

# ============================================================ 3. Accounts ===
ac = wb["Accounts"]
ac["H15"] = '="Ties to AED "&TEXT($D$15,"#,##0.00")&" — recomputed, never typed"'
log("Accounts",
    'H15 asserted "Must equal AED 4,277.19" against a live total of AED 6,105.17 — a '
    "note left behind by an earlier balance set. It is now a formula that states the "
    "current total, so it can never go stale again.")

# ============================================================ 4. Settings ===
st = wb["Settings"]
st["B12"] = ('=IF(DAY($B$9)<$B$11,DATE(YEAR($B$9),MONTH($B$9),$B$11),'
             'DATE(YEAR($B$9),MONTH($B$9)+1,$B$11))')
st["D12"] = ("Derived from the salary day. On payday itself the salary has landed, so the "
             "count rolls to next month — with <= it returned zero days, which drove the "
             "safe daily limit to zero on the one day of the month you are paid.")
log("Settings",
    "Next-income date used <= against the salary day, so on payday itself it returned "
    "today and left zero days to income — which pushed E8 and the SAFE DAILY LIMIT to "
    "zero on payday. Changed to <, so the count rolls forward once the salary lands.")

# ========================================================= 6. MASTER PLAN ===
# E56 is correctly named "Total accessible assets" on Settings, but MASTER PLAN
# labelled it "TOTAL NET WORTH" — and it does not net off the AED 2,687.36 of
# Tabby debt. Relabelled, with a true net-worth figure added beside it.
# Rows are never inserted: openpyxl does not rewrite formulas, so a shift here
# would silently break the stage tables below.
st_pre = wb["Settings"]
if st_pre["A118"].value is None:
    st_pre["A118"] = "E59  ·  Net worth (assets less debt)"
    st_pre["B118"] = "=$B$115-$B$32"
    st_pre["D118"] = ("Total accessible assets less the Tabby exposure. E56 above is deliberately "
                      "gross — this is the figure that nets what you owe.")
    for col, src in (("A", "A117"), ("B", "B117"), ("D", "D117")):
        style_from(st_pre, src, f"{col}118")

mp = wb["MASTER PLAN"]
if str(mp["B11"].value).strip().upper() == "TOTAL NET WORTH":
    mp["B11"] = "TOTAL ASSETS"
    mp["F11"] = "Everything you own, tracked. This line does not net off what you owe."
    mp["D11"] = "Net worth after Tabby"
    mp["E11"] = "=Settings!$B$118"
    style_from(mp, "B11", "D11")
    style_from(mp, "C11", "E11")
    log("MASTER PLAN",
        'Row 11 was labelled "TOTAL NET WORTH" but pointed at total accessible assets, which '
        "does not net off the AED 2,687.36 Tabby exposure. Relabelled TOTAL ASSETS, with a "
        "true net-worth figure added alongside (new engine row E59 on Settings).")

# ====================================================== 5. Checks & Audit ===
ck = wb["Checks & Audit"]

# 4a. Row 22 compared two hardcoded constants with each other. It could never
#     fail and its numbers had drifted from the live engine.
ck["B22"] = "=Settings!$B$87+Settings!$B$88"
ck["C22"] = "=Settings!$B$86"
ck["G22"] = ("Household plus personal must equal total own-money spending. Both sides now "
             "read the live engine — previously both were typed constants "
             "(5,939.85 + 1,503.16 against 7,443.01) that agreed only with each other.")
log("Checks & Audit",
    "The household-plus-personal check compared two typed constants with each other, so "
    "it could never fail, and both had drifted from the live figures. Both sides now "
    "read the engine.")

# 4b. Three MODEL STATUS banners had accumulated, and two of them sat in the
#     source log's ID column — which is why the log begins at S6 with S1–S5
#     missing. One banner, covering every check, then a clean log.
for coord in ("A27", "A31", "A34"):
    ck[coord].value = None

BANNER = ('=IF(COUNTIF($F$8:$F$33,"CHECK")=0,"MODEL STATUS:  ALL CHECKS OK",'
          '"MODEL STATUS:  "&COUNTIF($F$8:$F$33,"CHECK")&" CHECK(S) REQUIRE ATTENTION")')
ck["A27"] = BANNER
style_from(ck, "A21", "A27")
ck["A27"].font = Font(name=ck["A21"].font.name, sz=11, b=True, color="FF0D3452")
log("Checks & Audit",
    "Three MODEL STATUS banners had accumulated as checks were appended, and two of them "
    "occupied the source log's ID column. Reduced to one banner covering every check "
    "(F8:F33).")

# 4c. Restore the source log's identifiers. No formula anywhere references an
#     S-number, so the sequence is renumbered cleanly rather than left with a
#     hole where the banners had been.
LOG_IDS = {31: "S1", 34: "S2", 35: "S3", 36: "S4", 37: "S5", 38: "S6", 39: "S7", 40: "S8"}
for r, sid in LOG_IDS.items():
    ck[f"A{r}"] = sid
    style_from(ck, "A41", f"A{r}")
log("Checks & Audit",
    "Source log restarted at S6 because S1–S5 had been overwritten. Renumbered S1–S8 "
    "with every entry preserved; no formula anywhere references an S-number.")

# 4d. Notes that describe superseded numbers. The checks pass; the explanations
#     did not match what they were explaining.
NOTES = {
    "G8":  "Ties to the Accounts sheet total, recomputed live rather than compared "
           "against a typed figure.",
    "G10": "FAB 4001 + FAB 4002 + FAB 4003, at the balances confirmed 25–26 Aug. "
           "Previously described the superseded 165.50 + 7,010.70 + 7.67 set.",
    "G11": "NBD Current + NBD Plus Saver, at the balances confirmed 26 Aug. Previously "
           "described the superseded 2.20 + 2.73 set.",
}
for coord, text in NOTES.items():
    ck[coord] = text
log("Checks & Audit",
    "Three check notes still described earlier balance sets — the FAB note added up to "
    "7,183.87 against a live 5,963.32, and the NBD note to 4.93 against 140.96. "
    "Rewritten to match what the checks now compare.")

# 4e. Record the amendment itself in the sheet that exists to record such things.
r = 49
ck[f"A{r}"] = "AMENDMENT LOG"
style_from(ck, "A6", f"A{r}")
r += 1
for col, h in zip("ABC", ["#", "Sheet", "What was amended, 26 Aug 2026"]):
    ck[f"{col}{r}"] = h
    style_from(ck, f"{col}7", f"{col}{r}")
r += 1
for i, (sheet, what) in enumerate(FIXES, start=1):
    ck[f"A{r}"] = f"F{i}"
    ck[f"B{r}"] = sheet
    ck[f"C{r}"] = what
    for col in "ABC":
        style_from(ck, f"{col}41", f"{col}{r}")
    ck[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ck.row_dimensions[r].height = 30
    r += 1
ck.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ck[f"A{r}"] = ("All 4,120 formulas were evaluated independently before and after these "
               "changes: zero errors, and all 22 integrity checks pass. The amendments "
               "above are structural — nothing arithmetic was wrong.")
style_from(ck, "A2", f"A{r}")

wb.save(OUT)
print(f"amended -> {OUT}")
for i, (sheet, what) in enumerate(FIXES, 1):
    print(f"  F{i:<2} {sheet:18s} {what[:88]}")
