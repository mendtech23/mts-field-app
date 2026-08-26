"""Static validation of every formula in the workbook, independent of
LibreOffice: reference resolution, function whitelist and cycle detection.

This does not prove the arithmetic is right — verify_numbers.py does that —
but it catches the failure modes that produce #REF!, #NAME? and hangs.
"""
import re, sys, collections
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

F = sys.argv[1] if len(sys.argv) > 1 else "work.xlsx"
wb = openpyxl.load_workbook(F)
names = set(wb.sheetnames)

# Functions LibreOffice evaluates without an _xlfn. prefix.
SAFE = {
    "SUM", "SUMIFS", "SUMIF", "SUMPRODUCT", "COUNT", "COUNTA", "COUNTIF", "COUNTIFS",
    "IF", "AND", "OR", "NOT", "MAX", "MIN", "ABS", "AVERAGE", "ROUND", "INT",
    "IFERROR", "INDEX", "MATCH", "TEXT", "YEAR", "MONTH", "DAY", "DATE", "TODAY",
    "MOD", "POWER", "SQRT", "LEN", "TRIM", "CONCATENATE", "VALUE",
}

CELL = re.compile(
    r"(?:'(?P<q>[^']+)'|(?P<b>[A-Za-z_][A-Za-z0-9_ .]*))?!?"
    r"(?P<ref>\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?)"
)
FUNC = re.compile(r"([A-Z][A-Z0-9._]*)\s*\(")

errors, warnings = [], []
graph = collections.defaultdict(set)
formula_cells = []


def cells_of(sheet, ref):
    mn_c, mn_r, mx_c, mx_r = range_boundaries(ref.replace("$", ""))
    for r in range(mn_r, mx_r + 1):
        for c in range(mn_c, mx_c + 1):
            yield (sheet, r, c)


for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            formula_cells.append((ws.title, cell.coordinate, v))
            body = v[1:]
            # strip string literals so text like "AED " never parses as a ref
            stripped = re.sub(r'"[^"]*"', '""', body)

            for fn in FUNC.findall(stripped):
                if fn not in SAFE:
                    errors.append(f"{ws.title}!{cell.coordinate}: unsupported function {fn}()")

            for m in CELL.finditer(stripped):
                sheet = m.group("q") or m.group("b") or ws.title
                if m.group("q") or m.group("b"):
                    if sheet not in names:
                        errors.append(f"{ws.title}!{cell.coordinate}: unknown sheet '{sheet}'")
                        continue
                ref = m.group("ref")
                try:
                    targets = list(cells_of(sheet, ref))
                except Exception as e:
                    errors.append(f"{ws.title}!{cell.coordinate}: bad reference {ref} ({e})")
                    continue
                if len(targets) == 1:
                    s, r, c = targets[0]
                    tgt = wb[s].cell(row=r, column=c)
                    if tgt.value is None:
                        warnings.append(
                            f"{ws.title}!{cell.coordinate} -> {s}!{get_column_letter(c)}{r} is empty")
                for t in targets:
                    graph[(ws.title, cell.row, cell.column)].add(t)

# ---- cycle detection over the reference graph -------------------------------
WHITE, GREY, BLACK = 0, 1, 2
colour = {}
cycles = []


def visit(node, stack):
    colour[node] = GREY
    stack.append(node)
    for nxt in graph.get(node, ()):
        c = colour.get(nxt, WHITE)
        if c == GREY:
            i = stack.index(nxt)
            cycles.append(stack[i:] + [nxt])
        elif c == WHITE and nxt in graph:
            visit(nxt, stack)
    stack.pop()
    colour[node] = BLACK


sys.setrecursionlimit(20000)
for n in list(graph):
    if colour.get(n, WHITE) == WHITE:
        visit(n, [])

fmt = lambda n: f"{n[0]}!{get_column_letter(n[2])}{n[1]}"
print(f"formulas checked : {len(formula_cells)}")
print(f"errors           : {len(errors)}")
print(f"cycles           : {len(cycles)}")
print(f"empty-ref warns  : {len(warnings)}")
for e in errors[:40]:
    print("  ERROR", e)
for c in cycles[:10]:
    print("  CYCLE", " -> ".join(fmt(n) for n in c))
for w in warnings[:40]:
    print("  warn ", w)
sys.exit(1 if errors or cycles else 0)
