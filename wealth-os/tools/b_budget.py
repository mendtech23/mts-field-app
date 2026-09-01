"""Step B — FX & Assumptions sheet, Recent Ledger helper columns,
Budget sheet and Spend Analysis sheet.
"""
import sys, datetime as dt
sys.path.insert(0, "build")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import style as S

F = "work.xlsx"
wb = openpyxl.load_workbook(F)

LEDGER = "'Recent Ledger'"
LR = f"{LEDGER}!$D$6:$D$101"      # amount
LI = f"{LEDGER}!$I$6:$I$101"      # budget category
LJ = f"{LEDGER}!$J$6:$J$101"      # counts as spend
LK = f"{LEDGER}!$K$6:$K$101"      # split
LC = f"{LEDGER}!$C$6:$C$101"      # merchant

# ======================================================= FX & ASSUMPTIONS ====
ws = wb.create_sheet("FX & Assumptions")
S.setup(ws)
S.widths(ws, {"A": 38, "B": 18, "C": 16, "D": 30, "E": 14, "F": 56})
S.title(ws, "Rates, Returns and Planning Assumptions", 6)
S.subtitle(ws, "Every lever the rest of the workbook depends on lives here. "
               "Blue figures on a yellow fill are the only cells meant to be edited by hand; "
               "black figures are calculated. Change a lever here and Budget, Net Worth, "
               "Wealth Plan, Goals and AI Advisor all move together.", 6)

def lever(row, label, value, unit, source, note, fmt=S.NUM4):
    S.put(ws, f"A{row}", label)
    S.put(ws, f"B{row}", value, fmt=fmt, color=S.INPUT_BLUE, fill=S.LEVER_FILL, bold=True)
    S.put(ws, f"C{row}", unit)
    S.put(ws, f"D{row}", source)
    S.put(ws, f"E{row}", "Lever")
    S.put(ws, f"F{row}", note, wrap=True)

def calc(row, label, formula, unit, source, note, fmt=S.NUM4):
    S.put(ws, f"A{row}", label)
    S.put(ws, f"B{row}", formula, fmt=fmt)
    S.put(ws, f"C{row}", unit)
    S.put(ws, f"D{row}", source)
    S.put(ws, f"E{row}", "Formula")
    S.put(ws, f"F{row}", note, wrap=True)

S.band(ws, 5, "CURRENCY", 6)
S.header(ws, 6, ["Assumption", "Value", "Unit", "Basis", "Type", "Note"])
lever(7, "AED per INR", 0.0385333, "AED / INR", "Derived from the last confirmed SIP transfer",
      "AED 462.40 funded INR 12,000 on 10 Aug 2026, so 462.40 / 12,000 = 0.0385333. "
      "Refresh this whenever a new transfer settles at a different rate.")
calc(8, "INR per AED", "=IF(B7=0,0,1/B7)", "INR / AED", "Reciprocal of the row above",
     "Roughly INR 25.95 per dirham at the last confirmed transfer rate.")
lever(9, "AED per USD", 3.6725, "AED / USD", "UAE dirham peg to the US dollar",
      "The dirham has been pegged at 3.6725 since 1997; treat as fixed unless the peg changes.")
calc(10, "USD per AED", "=IF(B9=0,0,1/B9)", "USD / AED", "Reciprocal of the row above",
     "Used to read the Amana statement back into dirhams.")

S.band(ws, 12, "RETURN, INFLATION AND RISK ASSUMPTIONS", 6)
S.header(ws, 13, ["Assumption", "Value", "Unit", "Basis", "Type", "Note"])
lever(14, "Indian equity mutual funds — expected return", 0.11, "p.a. nominal",
      "Long-run planning assumption", "Nominal INR return before currency effects. "
      "Deliberately below the trailing returns shown on the Investments sheet: a snapshot gain is "
      "not a forecast.", fmt=S.PCT)
lever(15, "Global equity (Amana) — expected return", 0.08, "p.a. nominal",
      "Long-run planning assumption",
      "Broad developed-market equity assumption in USD.", fmt=S.PCT)
lever(16, "Cash and savings — expected return", 0.02, "p.a. nominal",
      "UAE savings account range",
      "FAB and NBD balances earn close to nothing; 2% is generous.", fmt=S.PCT)
lever(17, "Silver / commodity — expected return", 0.05, "p.a. nominal",
      "Long-run planning assumption",
      "Applies to the Nippon Silver ETF FoF holding only.", fmt=S.PCT)
lever(18, "UAE inflation", 0.025, "p.a.", "Planning assumption",
      "Used to restate the Wealth Plan into today's money.", fmt=S.PCT)
lever(19, "Scenario adjustment to returns", 0.0, "p.a.", "Owner-set stress lever",
      "Set to 0 for the base case, -0.03 for a bear case, +0.02 for a bull case. "
      "Applies to every growth assumption in the Wealth Plan.", fmt=S.PCT)
calc(20, "Blended portfolio planning return",
     "=IF('Investments'!B50=0,B14,(('Investments'!B43+'Investments'!B44+'Investments'!B45"
     "+'Investments'!B46)*B14+'Investments'!B47*B17+'Investments'!B48*B15"
     "+'Investments'!B49*B16)/'Investments'!B50)+B19",
     "p.a. nominal", "Weighted by the asset-allocation table on Investments",
     "Each sleeve is weighted by its present value at its own expected return: Indian equity funds "
     "at the equity rate, the silver holding at the commodity rate, Amana at the global rate and "
     "idle broker cash at the cash rate. The scenario adjustment is then added.", fmt=S.PCT)
calc(21, "Real (after-inflation) planning return", "=(1+B20)/(1+B18)-1", "p.a. real",
     "Blended return deflated by inflation",
     "The number that actually matters for buying power.", fmt=S.PCT)

S.band(ws, 23, "INCOME, HOUSING AND SAVINGS ASSUMPTIONS", 6)
S.header(ws, 24, ["Assumption", "Value", "Unit", "Basis", "Type", "Note"])
calc(25, "Net monthly salary", "='Inputs'!B18", "AED", "Inputs sheet, July salary baseline",
     "Linked, not typed. Update the baseline on Inputs and every sheet follows.", fmt=S.AED)
lever(26, "Annual salary increment", 0.04, "p.a.", "Planning assumption",
      "Applied from the first full plan year onward.", fmt=S.PCT)
lever(27, "Rent cheques per year", 4, "cheques", "ASSUMPTION — CONFIRM AGAINST THE LEASE",
      "The lease schedule in this file shows one cheque of AED 11,750 due 22 Oct 2026 but not "
      "the cadence. Four cheques a year is the common Dubai arrangement and is what the monthly "
      "rent accrual on Budget uses. If the real schedule is different, change this one cell.",
      fmt="0")
calc(28, "Annual rent", "='Inputs'!B30*B27", "AED", "Cheque value times cheques per year",
     "Drives the monthly rent accrual used in the Budget.", fmt=S.AED)
calc(29, "Monthly rent accrual", "=B28/12", "AED", "Annual rent spread evenly",
     "Setting money aside every month is what turns the next cheque from a crisis into a transfer.",
     fmt=S.AED)
lever(30, "Target savings rate", 0.20, "of net income", "Wealth-building target",
      "Everything invested plus emergency-fund top-ups, as a share of net income.", fmt=S.PCT)
calc(31, "Current monthly SIP", "='Inputs'!B39", "AED", "Inputs sheet",
     "INR 12,000 at the last confirmed transfer rate.", fmt=S.AED)
lever(32, "Annual SIP step-up", 0.10, "p.a.", "Wealth-building lever",
      "Raising the SIP by 10% a year is the single highest-leverage habit in this plan; "
      "the Wealth Plan sensitivity block prices it.", fmt=S.PCT)
lever(33, "Emergency fund target", 6, "months of essential spend", "Standard planning rule",
      "Measured against the essential subtotal on the Budget sheet.", fmt="0")

S.band(ws, 35, "LEDGER WINDOW AND PLAN HORIZON", 6)
S.header(ws, 36, ["Assumption", "Value", "Unit", "Basis", "Type", "Note"])
calc(37, "Ledger first transaction", f"=MIN({LEDGER}!$A$6:$A$101)", "date",
     "Recent Ledger", "Earliest confirmed transaction in the ledger.", fmt=S.DATE)
calc(38, "Ledger last transaction", f"=MAX({LEDGER}!$A$6:$A$101)", "date",
     "Recent Ledger", "Latest confirmed transaction in the ledger.", fmt=S.DATE)
calc(39, "Days covered by the ledger", "=INT(B38)-INT(B37)+1", "days",
     "Inclusive day count", "The ledger is a partial month, so actuals are scaled to a "
     "monthly run rate before they are compared with the plan.", fmt="0")
calc(40, "Run-rate factor to a full month", "=IF(B39=0,0,30.44/B39)", "x",
     "Average month length divided by days covered",
     "Multiply any ledger actual by this to get a monthly run rate.", fmt=S.MULT)
lever(41, "Plan start", dt.datetime(2026, 9, 1), "date", "First full month of the plan",
      "The Wealth Plan projects forward from this date.", fmt=S.DATE)
lever(42, "Plan horizon", 20, "years", "Owner-set",
      "Length of the projection on the Wealth Plan sheet.", fmt="0")

S.note(ws, 44,
       "Assumption discipline: a lever is a number this workbook cannot prove. Every one of them is "
       "on this sheet, in blue on yellow, with its basis written next to it. Nothing else in the file "
       "hardcodes a rate or a return. The two levers most likely to be wrong are rent cheques per "
       "year and the AED/INR rate — confirm both against the lease and the next transfer receipt.", 6)
ws.row_dimensions[44].height = 46

# ================================================ RECENT LEDGER HELPERS =====
ws = wb["Recent Ledger"]
S.widths(ws, {"I": 22, "J": 15, "K": 15})
for r in (1, 2):
    for col in ("I", "J", "K"):
        ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=S.TITLE_BG)
for col in ("I", "J", "K"):
    ws[f"{col}3"].fill = PatternFill("solid", fgColor=S.HEAD_BG)
S.header(ws, 5, ["Budget category", "Counts as spend", "Split"], col0=9)

CATMAP = {
    "Household utilities": "Utilities & Telecom",
    "Lifestyle": "Lifestyle & Shopping",
    "Household other": "Family & Support",
    "Food & Dining": "Dining",
    "Dining": "Dining",
    "Groceries": "Groceries",
    "Travel & airfare": "Travel",
    "Transport": "Fuel & Transport",
    "Transport & Fuel": "Fuel & Transport",
    "Grooming": "Grooming",
    "Bank fee": "Bank Fees",
    "Reconciliation": "Unreconciled",
    "Transfer out": "Unreconciled",
}
NON_SPEND = {"Transfer", "Income", "Mixed inflow", "Refunded / excluded"}

for r in range(6, 102):
    cat = ws[f"E{r}"].value
    cls = ws[f"F{r}"].value
    if cls in NON_SPEND or cat not in CATMAP:
        S.put(ws, f"I{r}", "Excluded", color=S.MUTED)
        S.put(ws, f"J{r}", 0, fmt="0", color=S.MUTED)
        S.put(ws, f"K{r}", "Excluded", color=S.MUTED)
    else:
        S.put(ws, f"I{r}", CATMAP[cat])
        S.put(ws, f"J{r}", 1, fmt="0")
        S.put(ws, f"K{r}", "Personal" if cls == "Personal" else "Household")

ws["A3"].value = ("Personal and household transactions are separated; transfers and refunds never "
                  "count as spending. Columns I to K normalise every line into one budget category, "
                  "a 1/0 spend flag and a personal/household split so the Budget, Spend Analysis and "
                  "AI Advisor sheets can total it without any manual sorting.")

# ============================================================= BUDGET =======
ws = wb.create_sheet("Budget")
S.setup(ws)
S.widths(ws, {"A": 30, "B": 17, "C": 17, "D": 17, "E": 17, "F": 13, "G": 14, "H": 52})
S.title(ws, "Monthly Budget — Plan against Reality", 8)
S.subtitle(ws, "Column B is the plan. Column C is what the confirmed ledger actually shows for "
               "3–25 Aug 2026. Column D scales that partial month to a full-month run rate so the "
               "two are comparable. A red variance means the run rate is above plan.", 8)

def sumifs(cat):
    return f"=SUMIFS({LR},{LI},$A{{row}},{LJ},1)".replace("$A{row}", f"$A{cat}")

HDR = ["Line", "Monthly plan", "Ledger actual\n3–25 Aug", "Monthly run rate",
       "Variance vs plan", "% of income", "Priority", "Note"]

def line(r, label, plan, ledger_cat, priority, note, bold=False, fill=None):
    S.put(ws, f"A{r}", label, bold=bold, fill=fill)
    S.put(ws, f"B{r}", plan, fmt=S.AED, bold=bold, fill=fill)
    if ledger_cat:
        S.put(ws, f"C{r}", f"=SUMIFS({LR},{LI},$A{r},{LJ},1)", fmt=S.AED, bold=bold, fill=fill)
        S.put(ws, f"D{r}", f"=C{r}*'FX & Assumptions'!$B$40", fmt=S.AED, bold=bold, fill=fill)
        S.put(ws, f"E{r}", f"=D{r}-B{r}", fmt=S.AED, bold=bold, fill=fill)
    else:
        S.put(ws, f"C{r}", 0, fmt=S.AED, bold=bold, fill=fill, color=S.MUTED)
        S.put(ws, f"D{r}", f"=C{r}*'FX & Assumptions'!$B$40", fmt=S.AED, bold=bold, fill=fill)
        S.put(ws, f"E{r}", f"=D{r}-B{r}", fmt=S.AED, bold=bold, fill=fill)
    S.put(ws, f"F{r}", f"=IF($B$8=0,0,B{r}/$B$8)", fmt=S.PCT, bold=bold, fill=fill)
    S.put(ws, f"G{r}", priority, bold=bold, fill=fill)
    S.put(ws, f"H{r}", note, wrap=True, bold=bold, fill=fill)

S.band(ws, 5, "MONTHLY INCOME", 8)
S.header(ws, 6, HDR)
line(7, "Net salary", "='FX & Assumptions'!B25", None, "Core",
     "July baseline. The AED 2,906 expected on 26 Aug is the remainder of this month's cycle, "
     "not an extra month of pay.")
S.put(ws, "A8", "TOTAL MONTHLY INCOME", bold=True, fill=S.HEAD_BG)
S.put(ws, "B8", "=B7", fmt=S.AED, bold=True, fill=S.HEAD_BG)
for col in "CDEFG":
    S.put(ws, f"{col}8", None, fill=S.HEAD_BG)
S.put(ws, "C8", "=C7", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "D8", "=D7", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "F8", "=IF(B8=0,0,B8/B8)", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, "G8", "Core", bold=True, fill=S.HEAD_BG)
S.put(ws, "H8", "Side income from ticket dealing is treated as one-off recovery, not salary, "
                "so it never inflates the plan.", bold=True, fill=S.HEAD_BG, wrap=True)

S.band(ws, 10, "ESSENTIAL — THE BILLS THAT KEEP THE LIGHTS ON", 8)
S.header(ws, 11, HDR)
line(12, "Rent accrual", "='FX & Assumptions'!B29", None, "Critical",
     "The rent cheque is not a surprise, it is a subscription. Accruing it monthly is what removes "
     "the October panic permanently. No ledger line exists because nothing has been set aside "
     "monthly so far.")
line(13, "Utilities & Telecom", "='Inputs'!B33+'Inputs'!B34+'Inputs'!B35", True, "Critical",
     "DEWA AED 813.28 + du AED 590.98 + Etisalat AED 323.95. Two telecom lines for one person is "
     "the most obvious fixed-cost cut available.")
line(14, "Groceries", 450, True, "Essential",
     "Cooking at home is the cheapest lever that does not require earning more.")
line(15, "Fuel & Transport", 250, True, "Essential",
     "Fuel and Metro. Work-essential; protect this before cutting anything else.")
line(16, "Family & Support", 100, True, "Essential",
     "Medical and family support transfers.")
S.put(ws, "A17", "SUBTOTAL ESSENTIAL", bold=True, fill=S.HEAD_BG)
for col, f in (("B", "=SUM(B12:B16)"), ("C", "=SUM(C12:C16)"), ("D", "=SUM(D12:D16)"),
               ("E", "=SUM(E12:E16)")):
    S.put(ws, f"{col}17", f, fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "F17", "=IF($B$8=0,0,B17/$B$8)", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, "G17", "Critical", bold=True, fill=S.HEAD_BG)
S.put(ws, "H17", "This subtotal is the base of the emergency-fund target on the Goals sheet.",
      bold=True, fill=S.HEAD_BG, wrap=True)

S.band(ws, 19, "LIFESTYLE — THE PART THAT IS ACTUALLY A CHOICE", 8)
S.header(ws, 20, HDR)
line(21, "Dining", 200, True, "Discretionary",
     "The single largest controllable category in the ledger. Restaurant and delivery spending "
     "has been running far above any plan that also funds rent.")
line(22, "Lifestyle & Shopping", 100, True, "Discretionary",
     "Convenience-store and hotel spending. Frozen under the damage-control cap.")
line(23, "Grooming", 65, True, "Discretionary",
     "One barber visit a month.")
line(24, "Travel", 0, True, "Deferred",
     "The AED 1,430 Emirates ticket sits in the ledger with a refund pending. No further travel "
     "spend is planned before the October rent cheque clears.")
line(25, "Bank Fees", 10, True, "Avoidable",
     "Transfer fees. Avoidable with a little batching.")
line(26, "Unreconciled", 0, True, "Control",
     "Balance-derived movements with no confirmed merchant. Target is zero: every dirham should "
     "have a name.")
S.put(ws, "A27", "SUBTOTAL LIFESTYLE", bold=True, fill=S.HEAD_BG)
for col, f in (("B", "=SUM(B21:B26)"), ("C", "=SUM(C21:C26)"), ("D", "=SUM(D21:D26)"),
               ("E", "=SUM(E21:E26)")):
    S.put(ws, f"{col}27", f, fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "F27", "=IF($B$8=0,0,B27/$B$8)", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, "G27", "Discretionary", bold=True, fill=S.HEAD_BG)
S.put(ws, "H27", "Every dirham cut here is a dirham that can fund rent, kill Tabby or buy units.",
      bold=True, fill=S.HEAD_BG, wrap=True)

S.band(ws, 29, "DEBT AND WEALTH — WHAT BUILDS THE FUTURE", 8)
S.header(ws, 30, HDR)
line(31, "Tabby repayment", "='Inputs'!B36", None, "Critical",
     "The no-fee minimum due 3 Sep. Paying the full statement is better whenever rent and "
     "essential bills stay funded.")
line(32, "Nippon SIP", "='FX & Assumptions'!B31", None, "Wealth",
     "INR 12,000 a month. This is the only line in the budget that compounds.")
line(33, "Emergency fund top-up", 200, None, "Wealth",
     "The emergency fund holds AED 7.67 against a six-month target. AED 200 a month is the "
     "smallest amount that makes the number move.")
S.put(ws, "A34", "SUBTOTAL DEBT AND WEALTH", bold=True, fill=S.HEAD_BG)
for col, f in (("B", "=SUM(B31:B33)"), ("C", "=SUM(C31:C33)"), ("D", "=SUM(D31:D33)"),
               ("E", "=SUM(E31:E33)")):
    S.put(ws, f"{col}34", f, fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "F34", "=IF($B$8=0,0,B34/$B$8)", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, "G34", "Wealth", bold=True, fill=S.HEAD_BG)
S.put(ws, "H34", "Debt repayment is defence; the SIP and the emergency fund are offence.",
      bold=True, fill=S.HEAD_BG, wrap=True)

S.band(ws, 36, "RESULT", 8)
S.header(ws, 37, ["Result", "Monthly plan", "Ledger actual\n3–25 Aug", "Monthly run rate",
                  "Variance vs plan", "% of income", "Verdict", "Note"])
res = [
    (38, "Total outflow", "=B17+B27+B34", "=C17+C27+C34", "=D17+D27+D34", "=D38-B38",
     '=IF($B$8=0,0,B38/$B$8)',
     '=IF(B39>=0,"Plan balances","Plan does not balance")',
     "Every planned dirham out, including debt repayment and investing."),
    (39, "Surplus / (deficit)", "=B8-B38", "=C8-C38", "=D8-D38", "=D39-B39",
     '=IF($B$8=0,0,B39/$B$8)',
     '=IF(B39>=0,"OK","ACTION")',
     "Negative means the plan spends more than the salary. The gap has to be closed by cutting "
     "lifestyle, cutting a telecom line, or earning more — not by borrowing."),
]
for r, label, b, c, d, e, f, g, h in res:
    fill = S.HEAD_BG
    S.put(ws, f"A{r}", label, bold=True, fill=fill)
    S.put(ws, f"B{r}", b, fmt=S.AED, bold=True, fill=fill)
    S.put(ws, f"C{r}", c, fmt=S.AED, bold=True, fill=fill)
    S.put(ws, f"D{r}", d, fmt=S.AED, bold=True, fill=fill)
    S.put(ws, f"E{r}", e, fmt=S.AED, bold=True, fill=fill)
    S.put(ws, f"F{r}", f, fmt=S.PCT, bold=True, fill=fill)
    S.put(ws, f"G{r}", g, bold=True, fill=fill)
    S.put(ws, f"H{r}", h, wrap=True, bold=True, fill=fill)

S.band(ws, 40, "BUDGET HEALTH RATIOS", 8)
S.header(ws, 41, ["Ratio", "Value", "Benchmark", "Verdict", "", "", "", "Note"])
ratios = [
    (42, "Savings rate achieved", "=IF(B8=0,0,(B32+B33)/B8)",
     "='FX & Assumptions'!B30", '=IF(B42>=C42,"ON TARGET","BELOW TARGET")',
     "Investing plus emergency-fund top-ups as a share of net income, against the target lever."),
    (43, "Essential ratio", "=IF(B8=0,0,B17/B8)", 0.50,
     '=IF(B43<=C43,"OK","HIGH")',
     "Essentials above half of net income leaves almost no room to build wealth. Rent accrual is "
     "the dominant term."),
    (44, "Lifestyle ratio", "=IF(B8=0,0,B27/B8)", 0.15,
     '=IF(B44<=C44,"OK","HIGH")',
     "Discretionary spending as a share of net income."),
    (45, "Run-rate lifestyle ratio (actual)", "=IF(B8=0,0,D27/B8)", 0.15,
     '=IF(B45<=C45,"OK","HIGH")',
     "The same ratio measured on what the ledger actually shows, not on the plan. "
     "This is the honest number."),
]
for r, label, val, bench, verdict, note in ratios:
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", val, fmt=S.PCT)
    S.put(ws, f"C{r}", bench, fmt=S.PCT, color=S.INPUT_BLUE if isinstance(bench, float) else None)
    S.put(ws, f"D{r}", verdict, bold=True)
    S.put(ws, f"H{r}", note, wrap=True)

S.note(ws, 47,
       "How to use this sheet: fix column B once a month, then let column C fill itself from the "
       "ledger. The only line that must never be cut is the rent accrual — everything else is "
       "negotiable, and the Budget only balances once the negotiation actually happens.", 8)
ws.row_dimensions[47].height = 40

# ==================================================== SPEND ANALYSIS ========
ws = wb.create_sheet("Spend Analysis")
S.setup(ws)
S.widths(ws, {"A": 30, "B": 17, "C": 13, "D": 17, "E": 17, "F": 15, "G": 52})
S.title(ws, "Where the Money Actually Went", 7)
S.subtitle(ws, "Built entirely from the confirmed ledger, 3–25 Aug 2026. Transfers between own "
               "accounts, salary credits and fully refunded charges are excluded, so nothing is "
               "counted twice and nothing is counted that never left.", 7)

CATS = ["Travel", "Utilities & Telecom", "Dining", "Lifestyle & Shopping", "Fuel & Transport",
        "Groceries", "Family & Support", "Grooming", "Unreconciled", "Bank Fees"]

S.band(ws, 5, "SPENDING BY CATEGORY", 7)
S.header(ws, 6, ["Category", "Total", "Share", "Personal", "Household", "Per day", "Read"])
READ = {
    "Travel": "One Emirates ticket with a refund pending. Strip it out and the underlying burn "
              "rate is far lower — but the cash still left the account.",
    "Utilities & Telecom": "A single DubaiPay payment covering du and Etisalat. Two telecom "
                           "contracts for one person is a standing, cancellable cost.",
    "Dining": "Restaurants, delivery and cafés. The largest genuinely controllable number in the "
              "file and the fastest route to closing the September gap.",
    "Lifestyle & Shopping": "Convenience stores and hotel bills. Small tickets, high frequency.",
    "Fuel & Transport": "Fuel and Metro. Work-essential; cut this last.",
    "Groceries": "Many small top-up trips rather than one weekly shop, which reliably costs more.",
    "Family & Support": "Medical support transfer. The duplicate was refunded and is excluded.",
    "Grooming": "One barber visit.",
    "Unreconciled": "Movements proven by the balance trail but with no confirmed merchant, plus "
                    "one outbound transfer to an untracked account. Target: zero.",
    "Bank Fees": "Transfer fee. Small, but pure leakage.",
}
r = 7
for cat in CATS:
    S.put(ws, f"A{r}", cat)
    S.put(ws, f"B{r}", f"=SUMIFS({LR},{LI},$A{r},{LJ},1)", fmt=S.AED)
    S.put(ws, f"C{r}", f"=IF($B${7+len(CATS)}=0,0,B{r}/$B${7+len(CATS)})", fmt=S.PCT)
    S.put(ws, f"D{r}", f'=SUMIFS({LR},{LI},$A{r},{LJ},1,{LK},"Personal")', fmt=S.AED)
    S.put(ws, f"E{r}", f'=SUMIFS({LR},{LI},$A{r},{LJ},1,{LK},"Household")', fmt=S.AED)
    S.put(ws, f"F{r}", f"=IF('FX & Assumptions'!$B$39=0,0,B{r}/'FX & Assumptions'!$B$39)", fmt=S.AED)
    S.put(ws, f"G{r}", READ[cat], wrap=True)
    r += 1
tot = r
S.put(ws, f"A{tot}", "TOTAL CONFIRMED SPENDING", bold=True, fill=S.HEAD_BG)
S.put(ws, f"B{tot}", f"=SUM(B7:B{tot-1})", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"C{tot}", f"=IF(B{tot}=0,0,B{tot}/B{tot})", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, f"D{tot}", f"=SUM(D7:D{tot-1})", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"E{tot}", f"=SUM(E7:E{tot-1})", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"F{tot}", f"=IF('FX & Assumptions'!$B$39=0,0,B{tot}/'FX & Assumptions'!$B$39)",
      fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, f"G{tot}", "Cross-check: this total must equal the sum of the 1-flagged rows in the "
                     "ledger. The Checks sheet proves it.", wrap=True, bold=True, fill=S.HEAD_BG)

MERCH = ["Emirates ticket", "du + Etisalat via DubaiPay", "RTS Business Bay Hotel",
         "Asas Al Madina General", "25 Hours F and B", "Emarat 1691 Mutina S",
         "Salkara / Team Taste", "e& Money medical support", "Cielo Kabab Restaurant",
         "Zomato", "ISK Gents Saloon", "Sahil Zam Zam Mandi"]
b2 = tot + 2
S.band(ws, b2, "TOP MERCHANTS BY VALUE", 7)
S.header(ws, b2 + 1, ["Merchant", "Total", "Share of spend", "Times seen", "Average ticket",
                      "Per day", "Read"])
r = b2 + 2
for m in MERCH:
    S.put(ws, f"A{r}", m)
    S.put(ws, f"B{r}", f'=SUMIFS({LR},{LC},$A{r},{LJ},1)', fmt=S.AED)
    S.put(ws, f"C{r}", f"=IF($B${tot}=0,0,B{r}/$B${tot})", fmt=S.PCT)
    S.put(ws, f"D{r}", f'=COUNTIFS({LC},$A{r},{LJ},1)', fmt="0")
    S.put(ws, f"E{r}", f"=IF(D{r}=0,0,B{r}/D{r})", fmt=S.AED)
    S.put(ws, f"F{r}", f"=IF('FX & Assumptions'!$B$39=0,0,B{r}/'FX & Assumptions'!$B$39)", fmt=S.AED)
    r += 1
S.put(ws, f"G{b2+2}", "Single largest line in the ledger; refund pending.", wrap=True)
S.put(ws, f"G{b2+3}", "One payment, two contracts. Consolidating to one line is a permanent saving.",
      wrap=True)
S.put(ws, f"G{b2+4}", "Three charges in one day. The kind of spend that only shows up when it is "
                      "totalled.", wrap=True)
S.put(ws, f"G{b2+5}", "The most frequent merchant in the file. Small tickets, constant repetition.",
      wrap=True)

b3 = r + 1
S.band(ws, b3, "BURN RATE AND CONTROL", 7)
S.header(ws, b3 + 1, ["Metric", "Value", "Benchmark", "Status", "", "", "Read"])
burn = [
    ("Total spend in the ledger window", f"=B{tot}", None, None,
     "Everything that genuinely left the accounts between 3 and 25 Aug."),
    ("Spend excluding the refundable ticket", f"=B{tot}-B7", None, None,
     "The underlying number once the one-off Emirates ticket is set aside."),
    ("Spend excluding ticket and utilities", f"=B{tot}-B7-B8", None, None,
     "Day-to-day living cost only: food, fuel, shops, grooming."),
    ("Daily burn — day-to-day living", f"=IF('FX & Assumptions'!$B$39=0,0,(B{tot}-B7-B8)/'FX & Assumptions'!$B$39)",
     "='Inputs'!B41/57", '=IF(B{r}<=C{r},"WITHIN CAP","ABOVE CAP")',
     "This is the headline finding of the whole sheet. The damage-control plan caps living costs "
     "at AED 5 a day; the confirmed ledger shows a multiple of that. The cap is not being kept, "
     "and that gap is exactly why the rent cheque is at risk."),
    ("Personal share of spending", f"=IF(B{tot}=0,0,D{tot}/B{tot})", 0.35,
     '=IF(B{r}<=C{r},"OK","HIGH")',
     "Personal lifestyle spending against household and work-essential spending."),
    ("Household share of spending", f"=IF(B{tot}=0,0,E{tot}/B{tot})", 0.65, None,
     "The balance of the split."),
    ("Discretionary share (dining, lifestyle, grooming)", f"=IF(B{tot}=0,0,(B9+B10+B14)/B{tot})",
     0.25, '=IF(B{r}<=C{r},"OK","HIGH")',
     "The share of spending that can be cut this week without any structural change."),
    ("Monthly run rate at this pace", f"=B{tot}*'FX & Assumptions'!$B$40", "='FX & Assumptions'!B25",
     '=IF(B{r}<=C{r},"WITHIN SALARY","ABOVE SALARY")',
     "The partial ledger scaled to a full month, against net salary. If this reads ABOVE SALARY, "
     "the shortfall is structural, not bad luck."),
]
r = b3 + 2
for label, val, bench, verdict, read in burn:
    S.put(ws, f"A{r}", label)
    fmt = S.PCT if "share" in label.lower() else S.AED
    S.put(ws, f"B{r}", val, fmt=fmt)
    if bench is not None:
        S.put(ws, f"C{r}", bench, fmt=fmt,
              color=S.INPUT_BLUE if isinstance(bench, float) else None)
    if verdict:
        S.put(ws, f"D{r}", verdict.replace("{r}", str(r)), bold=True)
    S.put(ws, f"G{r}", read, wrap=True)
    r += 1

S.note(ws, r + 1,
       "Read this sheet before the Budget. The Budget says what should happen; this sheet says what "
       "did. Where they disagree, this sheet is right.", 7)
ws.row_dimensions[r + 1].height = 32

wb.save(F)
print("step B ok")
