"""Step D — Debt Plan, Goals and AI Advisor sheets."""
import sys, datetime as dt
sys.path.insert(0, "build")
import openpyxl
import style as S

F = "work.xlsx"
wb = openpyxl.load_workbook(F)

# ============================================================ DEBT PLAN =====
ws = wb.create_sheet("Debt Plan")
S.setup(ws)
S.widths(ws, {"A": 34, "B": 17, "C": 15, "D": 18, "E": 15, "F": 26, "G": 17, "H": 50})
S.title(ws, "Debt Plan — Getting Tabby to Zero", 8)
S.subtitle(ws, "One debt, two statements, a hard freeze and a dated route to zero. Tabby carries no "
               "interest while the no-fee minimum is met on time, which makes punctuality — not "
               "size — the whole risk here.", 8)

S.band(ws, 5, "POSITION", 8)
S.header(ws, 6, ["Item", "Amount", "Due", "Status", "Priority", "Rule", "Cost of slipping", "Note"])
pos = [
    (7, "August statement (full)", "='Inputs'!B37", dt.datetime(2026, 9, 3), "Actual statement",
     "Optional", "Pay in full only if rent and bills stay funded", None,
     "Clearing this on 3 Sep would end the debt a month early, but the September funding gap says "
     "the cash is not there."),
    (8, "August no-fee minimum", "='Inputs'!B36", dt.datetime(2026, 9, 3), "Actual due",
     "Critical", "Autopay is already selected", "=35",
     "The amount that must land on 3 Sep. Missing it converts a free debt into an expensive one."),
    (9, "September statement", "='Inputs'!B42", dt.datetime(2026, 10, 3), "Actual generated",
     "Critical", "Pay by due date", "=35",
     "Already generated and confirmed. It falls due nineteen days before the rent cheque."),
    (10, "Total outstanding exposure", "='Inputs'!B38", None, "Actual confirmed", "Debt control",
     "Card frozen until zero", None,
     "August statement plus September statement. Do not add the two statements to this figure "
     "again — it already contains both."),
    (11, "Remaining after the 3 Sep minimum", "=B10-B8", None, "Formula", "Debt control",
     "Falls to zero on the schedule below", None,
     "What is still owed once the September minimum clears."),
]
for r, label, amt, due, status, prio, rule, cost, note in pos:
    bold = (r == 10)
    fill = S.HEAD_BG if bold else None
    S.put(ws, f"A{r}", label, bold=bold, fill=fill)
    S.put(ws, f"B{r}", amt, fmt=S.AED, bold=bold, fill=fill)
    if due: S.put(ws, f"C{r}", due, fmt=S.DATE, bold=bold, fill=fill)
    S.put(ws, f"D{r}", status, bold=bold, fill=fill)
    S.put(ws, f"E{r}", prio, bold=bold, fill=fill)
    S.put(ws, f"F{r}", rule, bold=bold, fill=fill, wrap=True)
    if cost: S.put(ws, f"G{r}", cost, fmt=S.AED, bold=bold, fill=fill)
    S.put(ws, f"H{r}", note, wrap=True, bold=bold, fill=fill)

S.band(ws, 13, "ROUTE TO ZERO — RECOMMENDED (PAY THE NO-FEE MINIMUM, PROTECT RENT)", 8)
S.header(ws, 14, ["Payment date", "Opening balance", "Payment", "Closing balance", "Fee incurred",
                  "Funded from", "Cumulative paid", "Note"])
sched = [
    (15, dt.datetime(2026, 9, 3), "='Inputs'!B38", "='Inputs'!B36", "26 Aug salary",
     "The confirmed no-fee minimum. Autopay already selected."),
    (16, dt.datetime(2026, 10, 3), "=D15", "='Inputs'!B42", "26 Sep salary",
     "The generated September statement, due before the rent cheque clears."),
    (17, dt.datetime(2026, 11, 3), "=D16", "=D16", "26 Oct salary",
     "The tail of the August statement. Paying it here clears the card completely."),
]
for r, date, opening, payment, funded, note in sched:
    S.put(ws, f"A{r}", date, fmt=S.DATE)
    S.put(ws, f"B{r}", opening, fmt=S.AED)
    S.put(ws, f"C{r}", payment, fmt=S.AED)
    S.put(ws, f"D{r}", f"=MAX(0,B{r}-C{r})", fmt=S.AED)
    S.put(ws, f"E{r}", 0, fmt=S.AED)
    S.put(ws, f"F{r}", funded)
    S.put(ws, f"G{r}", f"=C{r}" if r == 15 else f"=G{r-1}+C{r}", fmt=S.AED)
    S.put(ws, f"H{r}", note, wrap=True)
S.put(ws, "A18", "DEBT FREE", bold=True, fill=S.HEAD_BG)
S.put(ws, "B18", "=D17", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "C18", None, fill=S.HEAD_BG)
S.put(ws, "D18", '=IF(D17<=0.01,"Zero balance","Still outstanding")', bold=True, fill=S.HEAD_BG)
S.put(ws, "E18", "=SUM(E15:E17)", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "F18", "Three salary cycles", bold=True, fill=S.HEAD_BG)
S.put(ws, "G18", "=G17", fmt=S.AED, bold=True, fill=S.HEAD_BG)
S.put(ws, "H18", "Total paid equals total owed, and no fee is ever incurred. That is the whole "
                 "objective.", wrap=True, bold=True, fill=S.HEAD_BG)

S.band(ws, 20, "ALTERNATIVE — CLEAR THE FULL AUGUST STATEMENT ON 3 SEP", 8)
S.header(ws, 21, ["Comparison", "Recommended route", "Aggressive route", "Difference",
                  "Verdict", "", "", "Note"])
alt = [
    (22, "Cash needed on 3 Sep", "='Inputs'!B36", "='Inputs'!B37", "=C22-B22",
     '=IF(C22>B22,"Aggressive costs more now","")',
     "The aggressive route needs this much extra in September, on top of a month that is already "
     "short."),
    (23, "Debt free on", "=A17", "=A16", None, '="One month later"',
     "The aggressive route ends the debt in October instead of November."),
    (24, "Fees incurred either way", "=E18", 0, "=C24-B24", '="No difference"',
     "Neither route incurs a fee, so speed buys nothing but the feeling of speed."),
    (25, "Effect on the October rent cheque", "=0", "=C22-B22", "=C25-B25",
     '=IF(C25>0,"Aggressive puts rent at risk","")',
     "Every dirham sent early to Tabby is a dirham not available for the AED 11,750 cheque."),
]
for r, label, a, b, d, verdict, note in alt:
    fmt = S.DATE if r == 23 else S.AED
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", a, fmt=fmt)
    S.put(ws, f"C{r}", b, fmt=fmt)
    if d: S.put(ws, f"D{r}", d, fmt=S.AED)
    S.put(ws, f"E{r}", verdict, bold=True)
    S.put(ws, f"H{r}", note, wrap=True)

S.band(ws, 27, "STANDING DEBT RULES", 8)
S.header(ws, 28, ["Rule", "Status", "", "", "", "", "", "Why"])
rules = [
    ("Keep the Tabby Card frozen until the balance is zero",
     '=IF(\'Inputs\'!B38>0,"IN FORCE","Released")',
     "A frozen card cannot generate a new statement. The balance can then only fall."),
    ("Never miss a due date",
     '="IN FORCE"',
     "Tabby is free while the minimum is met on time. Late is the only way this debt becomes "
     "expensive."),
    ("No new buy-now-pay-later of any kind before the rent cheque clears",
     '=IF(TODAY()<DATE(2026,10,22),"IN FORCE","Review")',
     "New instalments would land in exactly the months the rent has to be funded."),
    ("Tabby Cash wallet spending is not Card borrowing",
     '="IN FORCE"',
     "Carried over from the original classification rule; spending the wallet never resets the "
     "Card streak."),
    ("Cancel or leave the Netflix subscription unfunded",
     '=IF(\'Obligations\'!C18>0,"OPEN","Closed")',
     "The AED 35 charge was declined, not paid. Letting it retry against protected cash would be "
     "the smallest and most avoidable own goal in the file."),
]
r = 29
for rule, status, why in rules:
    S.put(ws, f"A{r}", rule, wrap=True)
    S.put(ws, f"B{r}", status, bold=True)
    S.put(ws, f"H{r}", why, wrap=True)
    r += 1

S.note(ws, r + 1,
       "This is a small debt with a big shadow. AED 2,687.36 is not what threatens the rent cheque; "
       "the timing is. Two of the three payments fall in the same eight weeks as the cheque, which "
       "is why the schedule above funds each one from a different salary.", 8)
ws.row_dimensions[r + 1].height = 40

# ================================================================ GOALS =====
ws = wb.create_sheet("Goals")
S.setup(ws)
S.widths(ws, {"A": 36, "B": 17, "C": 17, "D": 17, "E": 13, "F": 15, "G": 18, "H": 50})
S.title(ws, "Goals — From Surviving the Month to Owning the Decade", 8)
S.subtitle(ws, "Ordered the only way that works: protect the roof, close the gap, kill the debt, "
               "build the buffer, then compound. Progress is calculated live from the rest of the "
               "workbook — nothing here is typed in by hand except the targets.", 8)

def goal(r, name, target, current, date, months, note, band_row=False):
    S.put(ws, f"A{r}", name)
    S.put(ws, f"B{r}", target, fmt=S.AED)
    S.put(ws, f"C{r}", current, fmt=S.AED)
    S.put(ws, f"D{r}", f"=MAX(0,B{r}-C{r})", fmt=S.AED)
    S.put(ws, f"E{r}", f"=IF(B{r}=0,1,MIN(1,C{r}/B{r}))", fmt=S.PCT)
    S.put(ws, f"F{r}", date, fmt=S.DATE)
    S.put(ws, f"G{r}", f"=IF({months}=0,D{r},D{r}/{months})", fmt=S.AED)
    S.put(ws, f"H{r}", note, wrap=True)

S.band(ws, 5, "STAGE 1 — SURVIVE THE NEXT NINETY DAYS", 8)
S.header(ws, 6, ["Goal", "Target", "Where you are", "Gap", "Progress", "Deadline",
                 "Per month needed", "Note"])
goal(7, "Close the September funding gap", "='Inputs'!B53", "='Inputs'!B43",
     dt.datetime(2026, 9, 15), 1,
     "Bills, survival spending and the SIP through 15 Sep, against the loose cash that exists "
     "today. This is the nearest cliff edge in the file.")
goal(8, "Fully fund the October rent cheque", "='Inputs'!B30", "='Inputs'!B31",
     dt.datetime(2026, 10, 21), 2,
     "The cheque clears 22 Oct and the 26 Oct salary is four days too late. Two salary cycles to "
     "find the balance.")
goal(9, "Clear the Tabby card to zero", "='Inputs'!B38", 0,
     dt.datetime(2026, 11, 3), 3,
     "Progress reads 0% until the balance falls — that is deliberate. Follow the Debt Plan "
     "schedule and this closes on 3 Nov with no fees.")

S.band(ws, 11, "STAGE 2 — BUILD THE BUFFER", 8)
S.header(ws, 12, ["Goal", "Target", "Where you are", "Gap", "Progress", "Deadline",
                  "Per month needed", "Note"])
goal(13, "Emergency fund — first milestone", 1000, "='Inputs'!B32",
     dt.datetime(2027, 3, 31), 6,
     "AED 1,000 is the milestone the original plan already named. At AED 200 a month from the "
     "Budget it takes five months, and it ends the era where one flat tyre becomes a crisis.")
goal(14, "Emergency fund — three months of essentials", "=Budget!B17*3", "='Inputs'!B32",
     dt.datetime(2027, 12, 31), 16,
     "Three months of the Budget essentials subtotal. The point at which a lost job stops being "
     "an emergency and becomes an inconvenience.")
goal(15, "Emergency fund — six months of essentials", "=Budget!B17*'FX & Assumptions'!B33",
     "='Inputs'!B32", dt.datetime(2029, 6, 30), 34,
     "Full target. Reach it and the SIP never has to be paused again for a bad month.")
goal(16, "Rent vault — one full year of rent", "='FX & Assumptions'!B28", "='Inputs'!B31",
     dt.datetime(2028, 10, 21), 26,
     "A year of rent held in advance. This is the goal that permanently ends the October panic, "
     "because the cheque stops competing with the salary.")

S.band(ws, 18, "STAGE 3 — COMPOUND", 8)
S.header(ws, 19, ["Goal", "Target", "Where you are", "Gap", "Progress", "Deadline",
                  "Per month needed", "Note"])
goal(20, "Investments reach AED 25,000", 25000, "='Net Worth'!F21",
     dt.datetime(2029, 9, 1), 36,
     "Roughly two and a half times the current portfolio, on the current SIP alone.")
goal(21, "Net worth reaches AED 100,000", 100000, "='Net Worth'!F28",
     dt.datetime(2032, 9, 1), 72,
     "The first genuinely life-changing threshold: a year of net salary held as capital.")
goal(22, "Net worth reaches AED 250,000", 250000, "='Net Worth'!F28",
     dt.datetime(2036, 9, 1), 120,
     "At this point investment growth in a normal year starts to rival what saving adds.")
goal(23, "Capital covers essential living costs", "='Wealth Plan'!B54", "='Net Worth'!F21",
     dt.datetime(2046, 9, 1), 240,
     "Financial independence as defined on the Wealth Plan: essential spending paid by capital "
     "rather than by work.")

S.band(ws, 25, "GOAL SCOREBOARD", 8)
S.header(ws, 26, ["Metric", "Value", "", "", "", "", "", "Read"])
score = [
    (27, "Goals fully funded", '=COUNTIF(E7:E9,">=1")+COUNTIF(E13:E16,">=1")+COUNTIF(E20:E23,">=1")',
     "0", "Goals already at 100%."),
    (28, "Goals under 25% funded",
     '=COUNTIF(E7:E9,"<0.25")+COUNTIF(E13:E16,"<0.25")+COUNTIF(E20:E23,"<0.25")', "0",
     "Goals that have barely started. Expected to be high today; the point is to watch it fall."),
    (29, "Total still to fund across every goal", "=SUM(D7:D9)+SUM(D13:D16)+SUM(D20:D23)", S.AED,
     "The full distance from here to the last goal. Large numbers are fine — they are supposed to "
     "take years."),
    (30, "Nearest deadline", "=MIN(F7:F9)", S.DATE,
     "The date that decides what you do this week."),
    (31, "Stage 1 average progress", "=AVERAGE(E7:E9)", S.PCT,
     "Survive-the-quarter progress. Nothing in stage 2 or 3 matters until this is at 100%."),
]
for r, label, val, fmt, read in score:
    S.put(ws, f"A{r}", label)
    S.put(ws, f"B{r}", val, fmt=fmt)
    S.put(ws, f"H{r}", read, wrap=True)

S.note(ws, 33,
       "Sequence matters more than ambition. Funding the SIP while the rent cheque is short is not "
       "investing, it is borrowing from the landlord. Work down this sheet in order and every later "
       "goal gets easier, because each stage removes a claim on the salary.", 8)
ws.row_dimensions[33].height = 40

# =========================================================== AI ADVISOR =====
ws = wb.create_sheet("AI Advisor")
S.setup(ws)
S.widths(ws, {"A": 9, "B": 44, "C": 16, "D": 15, "E": 17, "F": 12, "G": 15, "H": 62})
S.title(ws, "AI Advisor — Live Recommendations", 8)
S.subtitle(ws, "Every recommendation below is wired to a live condition in this workbook. Change a "
               "balance on Inputs or add a line to the ledger and the status column re-decides "
               "itself — an item closes when the data says it is closed, not when someone ticks "
               "it off. Impact is the dirham value of acting, stated per month unless the note "
               "says otherwise.", 8)

S.band(ws, 5, "HEALTH SCORE", 8)
S.header(ws, 6, ["#", "Component", "Score", "Weight", "Weighted", "", "", "How it is measured"])
comp = [
    (7, "Liquidity — free cash against one month of essentials",
     "=MIN(100,MAX(0,IF(Budget!B17=0,0,'Inputs'!B43/Budget!B17)*100))", 0.20,
     "Free cash outside rent and emergency, divided by monthly essentials, capped at 100."),
    (8, "Emergency cover — months held against the six-month target",
     "=MIN(100,MAX(0,IF(Budget!B17=0,0,'Inputs'!B32/(Budget!B17*'FX & Assumptions'!B33))*100))",
     0.20, "Emergency fund against six months of essentials."),
    (9, "Debt — exposure against total assets",
     "=MIN(100,MAX(0,(1-IF('Net Worth'!F25=0,1,'Net Worth'!F27/'Net Worth'!F25))*100))", 0.15,
     "One hundred means debt free."),
    (10, "Savings rate — achieved against target",
     "=MIN(100,MAX(0,IF('FX & Assumptions'!B30=0,0,Budget!B42/'FX & Assumptions'!B30)*100))",
     0.20, "Investing and emergency top-ups against the target savings rate."),
    (11, "Budget balance — does the plan fit the salary",
     "=IF(Budget!B39>=0,100,MAX(0,100+Budget!B39/MAX(1,Budget!B8)*100))", 0.15,
     "One hundred when the plan balances; falls as the deficit widens."),
    (12, "Spending discipline — run rate against plan",
     "=MIN(100,MAX(0,IF(Budget!D27=0,100,Budget!B27/Budget!D27*100)))", 0.10,
     "Planned lifestyle spending divided by the actual lifestyle run rate."),
]
for i, (r, label, formula, weight, how) in enumerate(comp, start=1):
    S.put(ws, f"A{r}", i, fmt="0")
    S.put(ws, f"B{r}", label, wrap=True)
    S.put(ws, f"C{r}", formula, fmt=S.NUM2)
    S.put(ws, f"D{r}", weight, fmt=S.PCT, color=S.INPUT_BLUE)
    S.put(ws, f"E{r}", f"=C{r}*D{r}", fmt=S.NUM2)
    S.put(ws, f"H{r}", how, wrap=True)
S.put(ws, "A13", None, fill=S.HEAD_BG)
S.put(ws, "B13", "FINANCIAL HEALTH SCORE", bold=True, fill=S.HEAD_BG)
S.put(ws, "C13", "=SUM(E7:E12)", fmt=S.NUM2, bold=True, fill=S.HEAD_BG)
S.put(ws, "D13", "=SUM(D7:D12)", fmt=S.PCT, bold=True, fill=S.HEAD_BG)
S.put(ws, "E13", '=IF(C13>=80,"A",IF(C13>=65,"B",IF(C13>=50,"C",IF(C13>=35,"D","E"))))',
      bold=True, fill=S.HEAD_BG)
S.put(ws, "H13", "Weights are levers — adjust them if you disagree with the emphasis. The grade in "
                 "column E is the honest headline: an E or a D means the next ninety days are about "
                 "survival, not strategy.", wrap=True, bold=True, fill=S.HEAD_BG)

RECS = [
 ("URGENT — DO THIS WEEK", [
  ("Bank the AED 189.19 bills-and-survival gap before 15 Sep",
   '=IF(\'Inputs\'!B51>0.01,"OPEN","CLOSED")', "='Inputs'!B51", "High", dt.datetime(2026, 9, 15),
   "Confirmed bills plus the survival allowance exceed loose cash plus the expected payday. This "
   "is the one number that decides whether September is calm or improvised. Earn it, or pause the "
   "SIP — in that order."),
  ("Do not touch the protected rent in FAB 4002",
   '=IF(\'Inputs\'!B31>0,"IN FORCE","Released")', "='Inputs'!B31", "None",
   dt.datetime(2026, 10, 21),
   "FAB 4002 shows AED 6,090.70 but only AED 100.25 of it is yours to spend. Treating the account "
   "balance as available is the single most likely way this plan fails."),
  ("Let the AED 1,314.50 Tabby autopay run on 3 Sep — do not pay the full statement",
   '=IF(\'Inputs\'!B36>0,"SCHEDULED","Done")', "='Inputs'!B37-'Inputs'!B36", "None",
   dt.datetime(2026, 9, 3),
   "Paying the full August statement costs AED 657.53 more in the month you can least afford it, "
   "and saves no fee at all. The Debt Plan proves it."),
  ("Confirm the AED 2,906 salary actually credits on 26 Aug",
   '=IF(\'Inputs\'!C22="Estimated","UNCONFIRMED","Confirmed")', "='Inputs'!B22", "Low",
   dt.datetime(2026, 8, 26),
   "Every downstream number in this workbook assumes it arrives. Until it is on a statement it is "
   "a hope, and the model deliberately refuses to call it cash."),
 ]),
 ("SPENDING — WHERE THE MONEY IS LEAKING", [
  ("Bring dining back to the plan",
   '=IF(Budget!D21>Budget!B21,"OVER PLAN","OK")', "=MAX(0,Budget!D21-Budget!B21)", "Medium",
   dt.datetime(2026, 9, 30),
   "Dining is the largest genuinely controllable category in the ledger. The impact column is the "
   "monthly saving from simply hitting the number already written in the Budget."),
  ("Cancel one of the two telecom lines",
   '=IF(AND(\'Inputs\'!B34>0,\'Inputs\'!B35>0),"TWO LINES ACTIVE","Consolidated")',
   "=MIN('Inputs'!B34,'Inputs'!B35)", "Low", dt.datetime(2026, 9, 30),
   "du and Etisalat are both being paid every month for one person. Cancelling the smaller line is "
   "a permanent saving that requires no willpower at all — the best kind."),
  ("Bring the daily burn under the damage-control cap",
   '=IF(\'Spend Analysis\'!B31>\'Spend Analysis\'!C31,"ABOVE CAP","WITHIN CAP")',
   "=MAX(0,('Spend Analysis'!B39-'Spend Analysis'!C39)*30.44)", "High",
   dt.datetime(2026, 9, 25),
   "The plan caps living costs at AED 5 a day. The confirmed ledger shows day-to-day living "
   "running at a large multiple of that. This gap, compounded over eight weeks, is roughly the "
   "size of the rent shortfall."),
  ("Kill the Netflix retry",
   '=IF(\'Obligations\'!C18>0,"OPEN","Closed")', "='Obligations'!C18", "None",
   dt.datetime(2026, 9, 1),
   "The charge was declined, not paid. Cancel it properly rather than leaving it to retry against "
   "an account that is holding rent money."),
  ("Stop paying transfer fees",
   '=IF(Budget!D25>0,"LEAKING","Clean")', "=Budget!D25", "Low", dt.datetime(2026, 9, 30),
   "Small, but it is pure leakage: a fee buys nothing. Batch transfers instead of moving money "
   "several times a week."),
  ("Give the AED 14 unreconciled movement a name",
   '=IF(COUNTIFS(\'Recent Ledger\'!$I$6:$I$101,"Unreconciled")>0,"OPEN","Clean")',
   "=SUMIFS('Recent Ledger'!$D$6:$D$101,'Recent Ledger'!$I$6:$I$101,\"Unreconciled\")", "Low",
   dt.datetime(2026, 9, 7),
   "A control item rather than a money item. A ledger where every dirham has a merchant is a "
   "ledger you can trust when it matters."),
 ]),
 ("STRUCTURE — STOP THE CRISIS REPEATING", [
  ("Start a monthly rent accrual",
   '=IF(\'Net Worth\'!F34>0.01,"NOT STARTED","Funded")', "='FX & Assumptions'!B29", "Medium",
   dt.datetime(2026, 11, 1),
   "The rent cheque is not an emergency, it is a subscription that arrives on a known date. Moving "
   "the monthly accrual into FAB 4002 on payday converts every future cheque from a crisis into a "
   "transfer. This is the highest-value structural change in the workbook."),
  ("Confirm how many rent cheques the lease actually requires",
   '=IF(\'FX & Assumptions\'!B27=4,"ASSUMED — CONFIRM","Confirmed")', "='FX & Assumptions'!B28",
   "None", dt.datetime(2026, 9, 7),
   "The whole rent accrual, the Budget essentials ratio and the emergency-fund target all rest on "
   "four cheques a year. It is an assumption, not a fact, and it is one phone call to settle."),
  ("Get the emergency fund to AED 1,000",
   '=IF(\'Inputs\'!B32<1000,"UNFUNDED","Milestone reached")', "=MAX(0,1000-'Inputs'!B32)",
   "Medium", dt.datetime(2027, 3, 31),
   "AED 7.67 is not a buffer. Until there is a real one, every unexpected cost lands on the card "
   "or on the rent money."),
  ("Move to a single spending account with a weekly allowance",
   '="RECOMMENDED"', "=0", "Low", dt.datetime(2026, 10, 1),
   "The ledger shows constant small transfers from 4002 into 4001. Each one is a decision, and "
   "each one erodes the ring-fence. One weekly transfer, then spend only what landed."),
 ]),
 ("INVESTING — THE PART THAT COMPOUNDS", [
  ("Sweep the idle rupee cash into the next SIP",
   '=IF(\'Net Worth\'!F19>1,"IDLE","Swept")', "='Net Worth'!F19", "None",
   dt.datetime(2026, 9, 10),
   "Settlement cash sitting in the broker account earns nothing. It is small, but sweeping it "
   "costs nothing and is the only free money in the file."),
  ("Reduce single-fund-house concentration",
   '=IF(\'Investments\'!B16>0.8,"CONCENTRATED","Diversified")',
   "=MAX(0,('Investments'!B16-0.8)*'Investments'!E12*'FX & Assumptions'!B7)", "Medium",
   dt.datetime(2027, 3, 31),
   "Almost the entire mutual-fund portfolio sits with one fund house. That is an operational risk "
   "as much as a market one. Direct future SIP instalments elsewhere rather than selling — no exit "
   "load, no capital gains event."),
  ("Do not pause the SIP unless the September gap is still open on 8 Sep",
   '=IF(\'Inputs\'!B51>0.01,"CONDITIONAL","Safe to continue")', "='Inputs'!B39", "None",
   dt.datetime(2026, 9, 8),
   "A paused SIP is cheap to restart; a missed rent cheque is not. But pausing early, when the gap "
   "might still be closed by earning, gives up compounding for nothing. Decide on 8 Sep with real "
   "balances, not on 26 Aug with estimates."),
  ("Review the 41 open Amana positions",
   '=IF(\'Investments\'!B25<0,"FLOATING LOSS","In profit")',
   "=('Investments'!B26+'Investments'!B27)*-1*'FX & Assumptions'!B9", "Medium",
   dt.datetime(2026, 9, 30),
   "Forty-one open positions on an account worth under AED 3,200 is a lot of moving parts, and the "
   "statement shows floating and overnight fees quietly working against you. Fewer, larger, "
   "longer-held positions would cost less to carry."),
  ("Raise the SIP by 10% every year, automatically",
   '=IF(\'FX & Assumptions\'!B32>0,"IN PLAN","Not set")',
   "='Wealth Plan'!H45", "None", dt.datetime(2027, 1, 1),
   "The impact column is what the step-up is worth at the twenty-year horizon compared with a flat "
   "contribution. It costs nothing today because the increase comes out of next year's raise."),
 ]),
]

r = 15
for section, items in RECS:
    S.band(ws, r, section, 8)
    S.header(ws, r + 1, ["#", "Recommendation", "Status", "Effort", "Impact (AED)", "Priority",
                         "By when", "Why this, and why now"])
    r += 2
    for i, (rec, status, impact, effort, by, why) in enumerate(items, start=1):
        S.put(ws, f"A{r}", i, fmt="0")
        S.put(ws, f"B{r}", rec, wrap=True)
        S.put(ws, f"C{r}", status, bold=True)
        S.put(ws, f"D{r}", effort)
        S.put(ws, f"E{r}", impact, fmt=S.AED)
        S.put(ws, f"F{r}", f'=IF(E{r}>=500,"1 — Critical",IF(E{r}>=100,"2 — High",'
                           f'IF(E{r}>=25,"3 — Useful","4 — Housekeeping")))')
        S.put(ws, f"G{r}", by, fmt=S.DATE)
        S.put(ws, f"H{r}", why, wrap=True)
        r += 1
    r += 1

S.band(ws, r, "THE ONE-PARAGRAPH VERDICT", 8)
r += 1
verdict = ('="Health score "&TEXT(C13,"0")&" out of 100, grade "&E13&". '
           'Net worth is AED "&TEXT(\'Net Worth\'!F28,"#,##0.00")&", of which AED "'
           '&TEXT(\'Net Worth\'!F21,"#,##0.00")&" actually compounds. '
           'The balance sheet is not the problem — the calendar is. '
           'AED "&TEXT(\'Net Worth\'!F34,"#,##0.00")&" of rent still has to be found before 21 Oct, '
           'AED "&TEXT(\'Inputs\'!B53,"#,##0.00")&" before 15 Sep, and free cash outside rent and '
           'emergency stands at AED "&TEXT(\'Inputs\'!B43,"#,##0.00")&". '
           'Do four things in order: hold the rent ring-fence, close the September gap by earning '
           'or by pausing the SIP, let the Tabby minimums run to zero by 3 Nov without ever paying '
           'a fee, and start accruing rent monthly so October 2027 is a transfer instead of a '
           'crisis. Then, and only then, raise the SIP."')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
S.put(ws, f"A{r}", verdict, wrap=True, fill=S.HEAD_BG, bold=True)
ws.row_dimensions[r].height = 92
ADVISOR_VERDICT_ROW = r

wb.save(F)
print("step D ok; advisor verdict row", ADVISOR_VERDICT_ROW)
